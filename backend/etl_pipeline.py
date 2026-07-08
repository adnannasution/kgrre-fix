"""
ETL Pipeline — baca file Excel mentah → knowledge graph CSV → import ke PostgreSQL.
Dijalankan di background thread; browser bisa ditutup setelah upload selesai.
"""
from __future__ import annotations

import json
import re
import threading
import time
import uuid
from pathlib import Path
from typing import Any

import duckdb
import openpyxl

from .config import UPLOADS_DIR
from .importer import JOBS, ImportJob, _create_job, _run_import
from .scanner import scan_package

ETL_JOBS: dict[str, ImportJob] = {}
ETL_JOBS_LOCK = threading.Lock()


# ---------------------------------------------------------------------------
# Sheet domain detection — berbasis konten kolom (nama file & sheet bebas)
# ---------------------------------------------------------------------------

def _col_set(headers: list[str]) -> set[str]:
    """Normalisasi header kolom ke set lowercase underscore."""
    return {re.sub(r'[^a-z0-9]+', '_', h.strip().lower()).strip('_') for h in headers if h and str(h).strip()}


def _detect_domain_by_columns(headers: list[str]) -> str | None:
    """Deteksi domain dari nama kolom — tidak perlu nama file tertentu."""
    cols = _col_set(headers)

    def has(*names: str) -> bool:
        return any(c in cols for c in names)

    def has_sub(sub: str) -> bool:
        return any(sub in c for c in cols)

    # PLO — kolom sangat spesifik
    if has('nomor_ijin', 'no_ijin', 'permit_number', 'nama_plo', 'status_plo'):
        return 'plo'
    if has('date_expired', 'masa_berlaku') and has('kapasitas', 'cakupan', 'nama_plo', 'nomor_ijin'):
        return 'plo'

    # OA Issue — mi_pi sangat khas
    if has('mi_pi', 'mip'):
        return 'oa_issue'

    # ICU Issue — mitigation/permanent_solution sangat spesifik
    if has('mitigation', 'permanent_solution') or (has('icu_status', 'icu') and has('issue', 'deskripsi_issue')):
        return 'icu_issue'
    if has('target_closed', 'tanggal_target') and has('mitigation') and has('issue', 'deskripsi_issue'):
        return 'icu_issue'

    # ATG — tag_no_tangki/tag_no_atg/status_atg sangat spesifik
    if has('tag_no_tangki', 'tag_no_atg') or has('status_atg', 'status_interkoneksi_atg'):
        return 'atg'
    if has('atg_eksisting') or (has_sub('atg') and has('program_2024', 'prokja')):
        return 'atg_program'

    # Pipeline inspection — beda dari inspection biasa: ada thickness measurement
    if has('last_measured_thickness', 'rem_life_years') or (has('fluida_service') and has('from_location', 'to_location')):
        return 'pipeline_inspection'

    # Power & Steam — status_n0 harus dicek SEBELUM readiness (yang juga punya status_operation)
    if has('status_n0') and has('kapasitas_max', 'average_actual'):
        return 'power_steam'

    # Monitoring Operasi — type_limitasi sangat spesifik
    if has('type_limitasi_process', 'limitasi_alert_process') or has('target_sts', 'limitasi_alert_sts'):
        return 'monitoring_operasi'

    # PAF (Plant Availability Factor) — plan_unplan + target_realisasi sangat spesifik
    if has('plan_unplan') and has('target_realisasi', 'type2'):
        return 'paf'

    # Critical Equipment — corrective_action + mitigasi_action
    if has('corrective_action') and has('mitigasi_action'):
        return 'critical_equipment'

    # TKDN — kdn + nominal + persentase
    if has('kdn') and has('nominal') and has('persentase'):
        return 'tkdn'

    # Zero Clamp — tag_no_ln atau type_damage + type_perbaikan
    if has('tag_no_ln') or (has('type_damage') and has('type_perbaikan', 'tanggal_dipasang')):
        return 'zero_clamp'

    # Bad Actor — category_action_plan + problem + tag_number/equipment
    if has('category_action_plan') and has('problem') and has('tag_number', 'tag_no', 'equipment'):
        return 'bad_actor'

    # Rotor — readiness_rotor / status_readiness_spare_rotor sangat spesifik
    if has('readiness_rotor', 'status_readiness_spare_rotor', 'spare_rotor'):
        return 'rotor'

    # Readiness subtypes — cek SEBELUM readiness umum (kolom status_operation sama)
    if has('status_operation', 'status_operasi', 'status_item') and \
            has('tag_no_tangki', 'level_oil', 'nama_tangki', 'no_tangki', 'kapasitas_tangki'):
        return 'readiness_tank'
    if has('status_operation', 'status_operasi', 'status_item') and \
            has('dermaga', 'nama_dermaga', 'jetty', 'nama_jetty', 'jenis_dermaga'):
        return 'readiness_jetty'
    if has('status_operation', 'status_operasi', 'status_item') and \
            has('spm', 'single_point_mooring', 'mooring', 'nama_spm'):
        return 'readiness_spm'

    # Workplan types — program_kerja dengan target/progres
    if has('progres_spm', 'target_spm', 'unit_spm') or (has_sub('spm') and has('program_kerja', 'progres', 'target')):
        return 'spm_workplan'
    if has('progres_tank', 'target_tank', 'unit_tank') or (has_sub('tangki') and has('program_kerja', 'progres', 'target')):
        return 'tank_workplan'
    if has('progres_jetty', 'target_jetty', 'unit_jetty') or (has_sub('jetty') and has('program_kerja', 'progres', 'target')):
        return 'jetty_workplan'

    # Readiness — status_operation sangat khas dan tidak ada di domain lain
    if has('status_operation', 'status_operasi'):
        return 'readiness'
    if has('status_item', 'rtl') and has('period_date', 'month_update'):
        return 'readiness'

    # OA Availability
    if has('actual_target') and has('value_perc', 'month_update', 'bulan'):
        return 'oa_availability'
    if has('operational_availability') or (has('oa_value', 'oa_target', 'oa_actual') and has('value_perc', 'actual_target')):
        return 'oa_availability'

    # Reliability — mtbf/mttr/running hours sangat spesifik
    if has('mtbf', 'mttr', 'reliability_index'):
        return 'reliability'
    if has('running_hours', 'jam_operasi', 'run_hours', 'operating_hours'):
        return 'reliability'
    if has('failure_date', 'failure_mode', 'breakdown_date') and has('equipment', 'tag_number', 'tag_no'):
        return 'reliability'

    # RKAP — cost_program/cost_element spesifik
    # cost_center saja tidak cukup — equipment master SAP juga punya kolom ini;
    # butuh minimal satu sinyal RKAP yang lebih spesifik
    if has('cost_program', 'cost_element') or (has('cost_center') and not has(
            'functional_loc', 'functional_location', 'floc',
            'catalog_profile', 'equipment_group', 'object_type',
            'description_of_technical_object', 'criticallity', 'criticality',
            'maintplant', 'maint_plant', 'planning_plant')):
        return 'rkap'
    if has_sub('rkap') or has_sub('irkap'):
        return 'rkap'
    if has('plan_idr', 'actual_idr', 'budget_idr') or (has('plan', 'actual', 'realisasi') and has('program', 'kegiatan')):
        return 'rkap'

    # Inspection plan
    if has_sub('inspection') and has('equipment', 'tag_number', 'tag_no', 'functional_loc', 'functional_location'):
        return 'inspection'
    if has('next_inspection', 'last_inspection', 'inspection_date', 'due_inspection'):
        return 'inspection'

    # RCPS — root cause
    if has('root_cause', 'akar_masalah') or (has('recommendation', 'rekomendasi') and has('finding', 'temuan')):
        return 'rcps_recommendation'
    if has_sub('rcps') or has('fishbone', 'why_1', 'why_2', 'why_3'):
        return 'rcps'

    # Org issue (sederhana) — issue list tanpa mitigation detail
    if has('issue', 'deskripsi_issue', 'paf') and has('responsible', 'pic', 'penanggung_jawab', 'priority'):
        return 'org_issue'

    # SAP Notification (terpisah dari Work Order) — qmnum adalah kunci SAP notifikasi
    if has('qmnum') or has('notifictn_type') or \
            (has('notification_no', 'notif_no') and not has('aufnr', 'order_no', 'maint_order')):
        return 'notification'

    # SAP Work Order — aufnr/order_no/order + work_center sinyal kuat
    if has('aufnr') or (
        has('order_no', 'maint_order', 'order') and
        has('work_center', 'main_work_center', 'main_workcenter', 'arbpl', 'main_workctr') and
        not has('qmnum', 'notifictn_type')  # jangan ambil notif
    ):
        return 'work_order'

    # Equipment Master vs Maintenance Order — paling ambigu
    eq_score = 0
    mo_score = 0

    # Equipment master signals
    if has('functional_loc', 'functional_location', 'floc'):          eq_score += 4
    if has('maintplant', 'planning_plant', 'maint_plant'):            eq_score += 3
    if has('catalog_profile', 'equipment_group', 'object_type'):      eq_score += 3
    if has('criticallity', 'criticality', 'critical_rank'):           eq_score += 3
    if has('description_of_technical_object'):                         eq_score += 3
    if has('equipment', 'tag_number', 'tag_no', 'tag_num'):           eq_score += 2
    if has('plant_area', 'location', 'plant_section'):                eq_score += 1

    # Maintenance order signals
    if has('order', 'order_number', 'aufnr', 'maint_order', 'order_no'): mo_score += 4
    if has('work_center', 'main_workcenter', 'arbpl', 'workcenter', 'main_workctr'):  mo_score += 4
    if has('order_type', 'auart', 'order_category'):                  mo_score += 3
    if has('notification', 'notification_no', 'qmnum'):               mo_score += 3
    if has('system_status', 'user_status', 'sttxt'):                  mo_score += 3
    if has('actual_start', 'actual_finish', 'basic_start', 'basic_finish', 'gstrp', 'gltrp'): mo_score += 2
    if has('equipment', 'tag_number', 'tag_no'):                      mo_score += 1

    if eq_score >= 5 and eq_score > mo_score:
        return 'equipment'
    if mo_score >= 5 and mo_score > eq_score:
        return 'work_order'
    if eq_score >= 4 and eq_score > mo_score + 2:
        return 'equipment'
    if mo_score >= 4 and mo_score > eq_score + 2:
        return 'work_order'

    return None


def _detect_domain(filename: str, sheet: str, headers: list[str] | None = None) -> str | None:
    """Deteksi domain sheet. Kolom adalah sinyal utama; nama file/sheet hanya fallback."""
    # 1. Deteksi dari konten kolom (paling reliable — bebas nama file)
    if headers:
        by_col = _detect_domain_by_columns(headers)
        if by_col:
            return by_col

    # 2. Fallback ke nama file/sheet untuk kasus ambigu atau kolom kosong
    stem = Path(filename).stem.lower()
    sheet_l = sheet.lower()
    k = f"{stem} {sheet_l}"

    if "all_ru_equipment" in stem:                                     return "equipment"
    if stem.startswith(("pt02_", "pt03_")):                           return "maintenance"
    if "rotor" in k:                                                   return "rotor"
    if any(x in k for x in ("work_order", "workorder", "sap_wo", "wo_sap")) or \
            (stem.startswith("wo_") or "_wo_" in stem):              return "work_order"
    if any(x in k for x in ("notif_sap", "sap_notif", "notifikasi_sap", "sap_notification")) or \
            (stem.startswith("notif") and "order" not in stem):      return "notification"
    if any(x in k for x in ("vw_reportirkapplanactual", "cost_program")) \
            or (any(x in k for x in ("rkap", "irkap")) and "alias_map" not in stem):
        return "rkap"
    if stem.startswith(("running_hours_", "n_0_")):                   return "reliability"
    if "inspection_plan" in k:                                         return "inspection"
    if any(x in k for x in ("apr_", "readiness_atg", "power_steam", "weekly_monitoring")):
        return "readiness"
    if any(x in k for x in ("workplan_spm", "spm_workplan", "prokja_spm", "program_kerja_spm")):   return "spm_workplan"
    if any(x in k for x in ("workplan_tank", "tank_workplan", "prokja_tank", "program_kerja_tank", "program_kerja_tangki")): return "tank_workplan"
    if any(x in k for x in ("workplan_jetty", "jetty_workplan", "prokja_jetty", "program_kerja_jetty")): return "jetty_workplan"
    if any(x in k for x in ("readiness_jetty", "jetty_readiness", "apr_jetty", "apr jetty")):     return "readiness_jetty"
    if any(x in k for x in ("readiness_spm", "spm_readiness", "apr_spm", "apr spm")):             return "readiness_spm"
    if any(x in k for x in ("readiness_tank", "tank_readiness", "readiness_tangki", "apr_tangki", "apr tangki", "apr_tank")): return "readiness_tank"
    if any(x in k for x in ("atg_monitoring", "monitoring_atg")):                                 return "atg"
    if any(x in k for x in ("program_kerja_atg", "atg_program", "prokja_atg")):                   return "atg_program"
    if "paf_issue" in k or ("paf" in stem and any(x in sheet_l for x in ("issue", "permasalahan"))): return "paf_issue"
    if any(x in k for x in ("issue_list",)):                          return "org_issue"
    if any(x in k for x in ("icu_database", "icu")):                 return "icu_issue"
    if "rcps" in stem:
        return "rcps_recommendation" if any(x in sheet_l for x in ("rekomendasi", "recommendation")) else "rcps"
    if "oa_data" in stem or "oa_" in stem or stem.startswith("oa ") or stem.startswith("oa data"):
        if any(x in sheet_l for x in ("allowance", "unplanned")):    return "oa_allowance"
        if any(x in sheet_l for x in ("issue", "permasalahan")):     return "oa_issue"
        return "oa_availability"
    if stem.startswith("plo") or "plo_" in stem:                      return "plo"
    return None


# ---------------------------------------------------------------------------
# Excel → DuckDB loader
# ---------------------------------------------------------------------------

def _read_sheet_rows(path: Path, sheet_name: str) -> list[tuple]:
    """Baca baris sheet. Coba read_only=True dulu; fallback ke normal jika hasilnya <= 1 baris."""
    for read_only in (True, False):
        wb = openpyxl.load_workbook(path, read_only=read_only, data_only=True)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows) > 1:
            return rows
        if not read_only:
            break
    return rows


def _load_excel_to_duckdb(con: duckdb.DuckDBPyConnection, path: Path) -> list[tuple[str, str, str, list[str]]]:
    """Baca semua sheet dari Excel, load ke DuckDB. Return list (table_name, filename, sheet, raw_headers)."""
    import pandas as pd
    loaded: list[tuple[str, str, str, list[str]]] = []

    wb_meta = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb_meta.sheetnames
    wb_meta.close()

    for sheet_name in sheet_names:
        rows = _read_sheet_rows(path, sheet_name)
        if not rows:
            continue
        raw_headers = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]

        # deduplikasi header untuk DuckDB
        seen: dict[str, int] = {}
        clean_headers: list[str] = []
        for h in raw_headers:
            h_clean = re.sub(r'[^a-zA-Z0-9_]', '_', h).lower() or "col"
            if h_clean in seen:
                seen[h_clean] += 1
                h_clean = f"{h_clean}_{seen[h_clean]}"
            else:
                seen[h_clean] = 0
            clean_headers.append(h_clean)

        data_rows = [
            tuple(str(v).strip() if v is not None else None for v in row)
            for row in rows[1:]
        ]
        if not data_rows:
            continue

        df = pd.DataFrame(data_rows, columns=clean_headers)
        df["_input_source_file"] = path.name
        df["_input_source_sheet"] = sheet_name
        df["_source_row"] = range(2, len(df) + 2)

        tname = f"src_{uuid.uuid4().hex[:8]}"
        con.register(tname, df)
        loaded.append((tname, path.name, sheet_name, raw_headers))

    return loaded


# ---------------------------------------------------------------------------
# DuckDB macros & reference tables
# ---------------------------------------------------------------------------

_MACROS_SQL = """
CREATE OR REPLACE MACRO norm_text(x) AS
    nullif(trim(regexp_replace(upper(coalesce(cast(x AS VARCHAR), '')),
                               '[^A-Z0-9]+', ' ', 'g')), '');

CREATE OR REPLACE MACRO norm_code(x) AS
    nullif(regexp_replace(upper(coalesce(cast(x AS VARCHAR), '')),
                          '[^A-Z0-9]+', '', 'g'), '');

CREATE OR REPLACE MACRO norm_equipment(x) AS norm_text(x);

-- Output RU di-alias dengan nama kilang ('RU II Dumai', dst) sesuai format data
-- yang diupload. Regex input tetap menerima 'RU II', '6201', maupun 'RU II Dumai'.
-- Catatan: _ru_key() di app.py menyaring balik ke 'RU II' (roman saja) sehingga
-- pencocokan Python tetap konsisten di kedua sisi.
CREATE OR REPLACE MACRO ru_normalize(x) AS (
    CASE
        WHEN norm_text(x) IN ('R201','R202','K201','K202','6201','6202')
          OR regexp_matches(norm_text(x), '(^| )RU ?(II|2)( |$)')
          OR regexp_matches(norm_text(x), '^RU2[A-Z]') THEN 'RU II Dumai'
        WHEN norm_text(x) IN ('R301','K301','6301')
          OR regexp_matches(norm_text(x), '(^| )RU ?(III|3)( |$)')
          OR regexp_matches(norm_text(x), '^RU3[A-Z]') THEN 'RU III Plaju'
        WHEN norm_text(x) IN ('R401','R402','R403','K401','6401')
          OR regexp_matches(norm_text(x), '(^| )RU ?(IV|4)( |$)')
          OR regexp_matches(norm_text(x), '^RU4[A-Z]') THEN 'RU IV Cilacap'
        WHEN norm_text(x) IN ('R501','K501','6501')
          OR regexp_matches(norm_text(x), '(^| )RU ?(V|5)( |$)')
          OR regexp_matches(norm_text(x), '^RU5[A-Z]') THEN 'RU V Balikpapan'
        WHEN norm_text(x) IN ('R601','K601','6601')
          OR regexp_matches(norm_text(x), '(^| )RU ?(VI|6)( |$)')
          OR regexp_matches(norm_text(x), '^RU6[A-Z]') THEN 'RU VI Balongan'
        WHEN norm_text(x) IN ('R701','K701','6701')
          OR regexp_matches(norm_text(x), '(^| )RU ?(VII|7)( |$)')
          OR regexp_matches(norm_text(x), '^RU7[A-Z]') THEN 'RU VII Kasim'
        ELSE NULL
    END
);

CREATE OR REPLACE MACRO ru_from_filename(x) AS (
    CASE
        WHEN regexp_matches(lower(coalesce(x,'')), 'ru[_ -]?ii([^iv]|$)') THEN 'RU II Dumai'
        WHEN regexp_matches(lower(coalesce(x,'')), 'ru[_ -]?iii') THEN 'RU III Plaju'
        WHEN regexp_matches(lower(coalesce(x,'')), 'ru[_ -]?iv([^i]|$)') THEN 'RU IV Cilacap'
        WHEN regexp_matches(lower(coalesce(x,'')), 'ru[_ -]?v([^i]|$)') THEN 'RU V Balikpapan'
        WHEN regexp_matches(lower(coalesce(x,'')), 'ru[_ -]?vi([^i]|$)') THEN 'RU VI Balongan'
        WHEN regexp_matches(lower(coalesce(x,'')), 'ru[_ -]?vii') THEN 'RU VII Kasim'
        ELSE NULL
    END
);
"""

_REFERENCE_SQL = """
CREATE TABLE IF NOT EXISTS ru_reference (
    refinery_unit VARCHAR, refinery_unit_id VARCHAR,
    display_order INTEGER, site_name VARCHAR
);
INSERT INTO ru_reference VALUES
    ('RU II Dumai',      'node_ru_' || md5('RU II Dumai'),      2, 'Dumai'),
    ('RU III Plaju',     'node_ru_' || md5('RU III Plaju'),     3, 'Plaju'),
    ('RU IV Cilacap',    'node_ru_' || md5('RU IV Cilacap'),    4, 'Cilacap'),
    ('RU V Balikpapan',  'node_ru_' || md5('RU V Balikpapan'),  5, 'Balikpapan'),
    ('RU VI Balongan',   'node_ru_' || md5('RU VI Balongan'),   6, 'Balongan'),
    ('RU VII Kasim',     'node_ru_' || md5('RU VII Kasim'),     7, 'Kasim');

CREATE TABLE IF NOT EXISTS plant_ru_map (plant VARCHAR, refinery_unit VARCHAR);
INSERT INTO plant_ru_map VALUES
    ('R201','RU II Dumai'),('R202','RU II Dumai'),('K201','RU II Dumai'),('K202','RU II Dumai'),
    ('6201','RU II Dumai'),('6202','RU II Dumai'),
    ('R301','RU III Plaju'),('K301','RU III Plaju'),('6301','RU III Plaju'),
    ('R401','RU IV Cilacap'),('R402','RU IV Cilacap'),('R403','RU IV Cilacap'),('K401','RU IV Cilacap'),('6401','RU IV Cilacap'),
    ('R501','RU V Balikpapan'),('K501','RU V Balikpapan'),('6501','RU V Balikpapan'),
    ('R601','RU VI Balongan'),('K601','RU VI Balongan'),('6601','RU VI Balongan'),
    ('R701','RU VII Kasim'),('K701','RU VII Kasim'),('6701','RU VII Kasim');

CREATE TABLE node_raw (
    node_id VARCHAR, node_type VARCHAR, business_key VARCHAR,
    label VARCHAR, domain VARCHAR, properties_json VARCHAR,
    source_file VARCHAR, source_sheet VARCHAR,
    source_row INTEGER, source_record_id VARCHAR
);

CREATE TABLE relationship_raw (
    source_node_id VARCHAR, target_node_id VARCHAR,
    relationship_type VARCHAR, domain VARCHAR,
    confidence DOUBLE, match_rule VARCHAR, is_candidate BOOLEAN,
    properties_json VARCHAR,
    source_file VARCHAR, source_sheet VARCHAR,
    source_row INTEGER, source_record_id VARCHAR
);

CREATE TABLE unmatched_raw (
    identifier VARCHAR, identifier_type VARCHAR,
    domain VARCHAR, source_file VARCHAR,
    source_sheet VARCHAR, source_row INTEGER, reason VARCHAR
);
"""


# ---------------------------------------------------------------------------
# Helper: safe column reference (hindari "column not found" di DuckDB)
# ---------------------------------------------------------------------------

def _union_cols(con: duckdb.DuckDBPyConnection, views: list[str]) -> set[str]:
    """Kumpulkan semua nama kolom yang ada di sekumpulan views."""
    cols: set[str] = set()
    for v in views:
        cols |= {r[0].lower() for r in con.execute(f'DESCRIBE {v}').fetchall()}
    return cols


def _cs(cols: set[str], *names: str, cast: bool = False, default: str = "''") -> str:
    """Build COALESCE hanya dari kolom yang benar-benar ada. Kolom tidak ada → default."""
    found = [f'cast("{n}" AS VARCHAR)' if cast else f'"{n}"' for n in names if n in cols]
    if not found:
        return default
    return f"coalesce({', '.join(found + [default])})"


# ---------------------------------------------------------------------------
# Node builders
# ---------------------------------------------------------------------------

def _build_ru_plant_nodes(con: duckdb.DuckDBPyConnection) -> None:
    con.execute("""
        INSERT INTO node_raw (node_id, node_type, business_key, label, domain, properties_json)
        SELECT refinery_unit_id, 'refinery_unit', refinery_unit, refinery_unit, 'asset',
               json_object('refinery_unit', refinery_unit, 'site_name', site_name,
                            'display_order', display_order)
        FROM ru_reference
    """)


def _build_equipment_nodes(con: duckdb.DuckDBPyConnection, views: list[str]) -> None:
    if not views:
        return
    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    eq_raw = _cs(c, 'equipment', 'tag_number', 'tag_no', default='NULL')
    plant_expr = f"norm_text({_cs(c, 'maintplant','planning_plant','maint_plant','plant', default='NULL')})"
    ru_expr = f"ru_normalize({_cs(c, 'refinery_unit','ru','refineryunit', default='NULL')})"
    con.execute(f"""
        CREATE TABLE equipment_master AS
        WITH base AS (
            SELECT
                norm_equipment({eq_raw}) AS equipment_code_normalized,
                {eq_raw} AS equipment_code_raw,
                regexp_replace(trim(cast(coalesce({eq_raw}, '') AS VARCHAR)), '/[0-9]+$', '') AS equipment_code_clean,
                {plant_expr} AS plant,
                coalesce(
                    {ru_expr},
                    (SELECT refinery_unit FROM plant_ru_map
                     WHERE plant = {plant_expr} LIMIT 1),
                    ru_from_filename(_input_source_file)
                ) AS refinery_unit,
                nullif(trim({_cs(c, 'functional_loc','functional_location','floc')}), '') AS functional_location,
                nullif(trim({_cs(c, 'catalog_profile','equipment_group','object_type')}), '') AS equipment_group,
                nullif(trim({_cs(c, 'equip_category','equipcategory','equipment_category')}), '') AS equip_category,
                nullif(trim({_cs(c, 'description_of_technical_object','description','equipment_name')}), '') AS description,
                nullif(trim({_cs(c, 'criticallity','criticality','critical_rank')}), '') AS criticallity,
                nullif(trim({_cs(c, 'location','plant_area','plant_section')}), '') AS plant_area,
                nullif(trim({_cs(c, 'manufacturer_of_asset','manufacturer','pabrikan')}), '') AS manufacturer,
                nullif(trim({_cs(c, 'model_type','model','type','tipe')}), '') AS model_type,
                nullif(trim({_cs(c, 'wbs_element','wbs')}), '') AS wbs_element,
                nullif(trim({_cs(c, 'cost_center')}), '') AS cost_center,
                nullif(trim({_cs(c, 'planner_group','planner')}), '') AS planner_group,
                nullif(trim({_cs(c, 'date_update_data','last_update','update_date')}), '') AS date_update_data,
                _input_source_file AS source_file,
                _input_source_sheet AS source_sheet,
                _source_row AS source_row,
                'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
            FROM ({union_sql})
            WHERE norm_equipment({eq_raw}) IS NOT NULL
        ),
        ranked AS (
            SELECT *, row_number() OVER (
                PARTITION BY coalesce(refinery_unit, 'UNKNOWN'), equipment_code_normalized
                ORDER BY try_cast(date_update_data AS DATE) DESC NULLS LAST,
                         ((functional_location IS NOT NULL)::INTEGER +
                          (description IS NOT NULL)::INTEGER +
                          (equipment_group IS NOT NULL)::INTEGER) DESC,
                         source_file, source_row DESC
            ) AS rn
            FROM base
        )
        SELECT 'node_equipment_' || md5(coalesce(refinery_unit, 'UNKNOWN') || '|' || equipment_code_normalized) AS equipment_id, *
        FROM ranked WHERE rn = 1
    """)

    con.execute("""
        INSERT INTO node_raw
        SELECT equipment_id, 'equipment',
               coalesce(refinery_unit, 'UNKNOWN') || '|' || equipment_code_normalized,
               coalesce(description, equipment_code_raw), 'asset',
               json_object(
                   'refinery_unit', refinery_unit,
                   'equipment_code_raw', equipment_code_raw,
                   'equipment_code_normalized', equipment_code_normalized,
                   'plant', plant,
                   'functional_location', functional_location,
                   'equipment_group', equipment_group,
                   'equip_category', equip_category,
                   'description', description,
                   'criticallity', criticallity,
                   'plant_area', plant_area,
                   'manufacturer', manufacturer,
                   'model_type', model_type,
                   'wbs_element', wbs_element,
                   'cost_center', cost_center,
                   'planner_group', planner_group,
                   'derived_ru_normalized', refinery_unit,
                   'derived_equipment_code_compact', norm_code(equipment_code_raw)
               ),
               source_file, source_sheet, source_row, source_record_id
        FROM equipment_master
    """)

    # Buat node Plant dan Functional Location dari equipment_master
    con.execute("""
        INSERT INTO node_raw (node_id, node_type, business_key, label, domain, properties_json)
        SELECT DISTINCT
            'node_plant_' || md5(coalesce(refinery_unit,'UNKNOWN') || '|' || plant),
            'plant', plant, plant, 'asset',
            json_object('plant', plant, 'refinery_unit', refinery_unit)
        FROM equipment_master WHERE plant IS NOT NULL
    """)
    con.execute("""
        INSERT INTO node_raw (node_id, node_type, business_key, label, domain, properties_json)
        SELECT DISTINCT
            'node_floc_' || md5(coalesce(refinery_unit,'UNKNOWN') || '|' || functional_location),
            'functional_location', functional_location, functional_location, 'asset',
            json_object('functional_location', functional_location, 'refinery_unit', refinery_unit, 'plant', plant)
        FROM equipment_master WHERE functional_location IS NOT NULL
    """)

    # RU → Plant → FLoc → Equipment hierarki
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json, source_file,
            source_sheet, source_row, source_record_id)
        SELECT r.refinery_unit_id,
               'node_plant_' || md5(coalesce(e.refinery_unit,'UNKNOWN') || '|' || e.plant),
               'REFINERY_UNIT_HAS_PLANT', 'asset', 1.0, 'plant_mapping', false,
               json_object('plant', e.plant), e.source_file, e.source_sheet, e.source_row, e.source_record_id
        FROM equipment_master e
        JOIN ru_reference r ON e.refinery_unit = r.refinery_unit
        WHERE e.plant IS NOT NULL
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json, source_file,
            source_sheet, source_row, source_record_id)
        SELECT
            'node_plant_' || md5(coalesce(e.refinery_unit,'UNKNOWN') || '|' || e.plant),
            'node_floc_' || md5(coalesce(e.refinery_unit,'UNKNOWN') || '|' || e.functional_location),
            'PLANT_HAS_FUNCTIONAL_LOCATION', 'asset', 1.0, 'floc_mapping', false,
            json_object('functional_location', e.functional_location),
            e.source_file, e.source_sheet, e.source_row, e.source_record_id
        FROM equipment_master e
        WHERE e.plant IS NOT NULL AND e.functional_location IS NOT NULL
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json, source_file,
            source_sheet, source_row, source_record_id)
        SELECT
            'node_floc_' || md5(coalesce(e.refinery_unit,'UNKNOWN') || '|' || e.functional_location),
            e.equipment_id,
            'FUNCTIONAL_LOCATION_HAS_EQUIPMENT', 'asset', 1.0, 'floc_mapping', false,
            json_object('equipment', e.equipment_code_raw),
            e.source_file, e.source_sheet, e.source_row, e.source_record_id
        FROM equipment_master e
        WHERE e.functional_location IS NOT NULL
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json, source_file,
            source_sheet, source_row, source_record_id)
        SELECT
            'node_plant_' || md5(coalesce(e.refinery_unit,'UNKNOWN') || '|' || e.plant),
            e.equipment_id,
            'PLANT_HAS_EQUIPMENT', 'asset', 1.0, 'plant_mapping', false,
            json_object('plant', e.plant),
            e.source_file, e.source_sheet, e.source_row, e.source_record_id
        FROM equipment_master e
        WHERE e.plant IS NOT NULL
    """)

    # RU → equipment edges
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json, source_file,
            source_sheet, source_row, source_record_id)
        SELECT r.refinery_unit_id, e.equipment_id, 'REFINERY_UNIT_HAS_EQUIPMENT',
               'asset', 1.0, 'plant_mapping', false,
               json_object('refinery_unit', e.refinery_unit),
               e.source_file, e.source_sheet, e.source_row, e.source_record_id
        FROM equipment_master e
        JOIN ru_reference r ON e.refinery_unit = r.refinery_unit
    """)


def _build_maintenance_nodes(con: duckdb.DuckDBPyConnection, views: list[str]) -> None:
    if not views:
        return
    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    ord_expr = _cs(c, 'order','order_no','order_number','aufnr','maint_order', cast=True, default='NULL')
    notif_expr = _cs(c, 'notification','notif','qmnum', cast=True, default='NULL')
    # Notifikasi SAP tanpa Order tetap dimasukkan dengan order_raw dari nomor notifikasi
    primary_id_expr = f"coalesce({ord_expr}, {notif_expr})"
    ru_expr = f"ru_normalize({_cs(c, 'refinery_unit','ru','plant','maintplant','refineryunit', default='NULL')})"
    con.execute(f"""
        CREATE TABLE maintenance_stage AS
        SELECT
            {ord_expr} AS order_raw,
            {notif_expr} AS notification_raw,
            coalesce(norm_code({ord_expr}), norm_code({notif_expr})) AS order_code,
            {_cs(c, 'description','kurztext','short_text','order_description')} AS order_desc,
            {_cs(c, 'order_type','auart','order_category','notifictn_type','notif_type')} AS order_type,
            {_cs(c, 'priority','priok')} AS priority,
            {_cs(c, 'user_status','txt04','ustatus')} AS user_status,
            {_cs(c, 'system_status','sttxt')} AS system_status,
            {_cs(c, 'reference_date','gstrp','basic_start','req_start','notif_date', cast=True)} AS reference_date,
            {_cs(c, 'total_planned_costs','geplk','planned_cost', cast=True, default="'0'")} AS planned_cost,
            {_cs(c, 'total_actual_costs','istko','actual_cost', cast=True, default="'0'")} AS actual_cost,
            {_cs(c, 'main_work_center','main_workcenter','arbpl','work_center','workcenter','main_workctr')} AS work_center,
            {_cs(c, 'equipment','tag_number','tag_no','equnr')} AS equipment_raw,
            coalesce({ru_expr}, ru_from_filename(_input_source_file)) AS refinery_unit,
            _input_source_file AS source_file,
            _input_source_sheet AS source_sheet,
            _source_row AS source_row,
            'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
        FROM ({union_sql})
        WHERE {primary_id_expr} IS NOT NULL
    """)

    # Tambah derived columns
    con.execute("""
        ALTER TABLE maintenance_stage ADD COLUMN order_id VARCHAR;
        UPDATE maintenance_stage
        SET order_id = 'node_order_' || md5(refinery_unit || '|' || order_code)
        WHERE order_code IS NOT NULL AND refinery_unit IS NOT NULL;

        ALTER TABLE maintenance_stage ADD COLUMN equipment_id VARCHAR;
        UPDATE maintenance_stage SET equipment_id = e.equipment_id
        FROM equipment_master e
        WHERE maintenance_stage.refinery_unit = e.refinery_unit
          AND norm_code(maintenance_stage.equipment_raw) = norm_code(e.equipment_code_clean)
          AND maintenance_stage.equipment_raw IS NOT NULL;
    """)

    con.execute("""
        INSERT INTO node_raw
        SELECT DISTINCT order_id, 'maintenance_order',
               refinery_unit || '|' || order_code,
               coalesce(nullif(order_desc,''), order_raw), 'maintenance',
               json_object(
                   'refinery_unit', refinery_unit,
                   'equipment_raw', equipment_raw,
                   'order_raw', order_raw,
                   'notification_raw', notification_raw,
                   'order_type', order_type,
                   'priority', priority,
                   'user_status', user_status,
                   'system_status', system_status,
                   'reference_date', reference_date,
                   'derived_planned_cost', planned_cost,
                   'derived_actual_cost', actual_cost,
                   'work_center', work_center,
                   'derived_is_open_order',
                       CASE WHEN user_status ILIKE '%WAMA%' OR user_status ILIKE '%WASR%'
                                 OR system_status ILIKE '%REL%' OR system_status ILIKE '%PCNF%'
                            THEN 'true' ELSE 'false' END,
                   'derived_status_bucket',
                       CASE WHEN user_status ILIKE '%WAMA%' THEN 'WAMA'
                            WHEN user_status ILIKE '%WASR%' THEN 'WASR'
                            WHEN system_status ILIKE '%TECO%' THEN 'TECO'
                            WHEN system_status ILIKE '%CLSD%' THEN 'CLSD'
                            ELSE 'OPEN' END
               ),
               source_file, source_sheet, source_row, source_record_id
        FROM maintenance_stage WHERE order_id IS NOT NULL
    """)

    # EQUIPMENT_HAS_MAINTENANCE_ORDER
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT equipment_id, order_id, 'EQUIPMENT_HAS_MAINTENANCE_ORDER',
               'maintenance', 1.0, 'ru_and_equipment_exact', false,
               json_object('match_token', equipment_raw, 'source_ru', refinery_unit),
               source_file, source_sheet, source_row, source_record_id
        FROM maintenance_stage WHERE equipment_id IS NOT NULL AND order_id IS NOT NULL
    """)

    # EQUIPMENT_HAS_NOTIFICATION (baris notifikasi tanpa order, atau notif_raw ada)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT equipment_id, order_id, 'EQUIPMENT_HAS_NOTIFICATION',
               'maintenance', 0.9, 'equipment_notification', false,
               json_object('notification', notification_raw, 'source_ru', refinery_unit),
               source_file, source_sheet, source_row, source_record_id
        FROM maintenance_stage
        WHERE equipment_id IS NOT NULL AND order_id IS NOT NULL
          AND notification_raw IS NOT NULL AND order_raw IS NULL
    """)

    # MAINTENANCE_ORDER_HAS_NOTIFICATION (order punya nomor notifikasi)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT o.order_id, n.order_id, 'MAINTENANCE_ORDER_HAS_NOTIFICATION',
               'maintenance', 1.0, 'order_notif_link', false,
               json_object('notification', n.notification_raw),
               n.source_file, n.source_sheet, n.source_row, n.source_record_id
        FROM maintenance_stage n
        JOIN maintenance_stage o
          ON norm_code(n.order_raw) = norm_code(o.notification_raw)
         AND n.refinery_unit = o.refinery_unit
        WHERE n.order_id IS NOT NULL AND o.order_id IS NOT NULL
          AND n.order_id != o.order_id
    """)

    # Unmatched equipment di maintenance
    con.execute("""
        INSERT INTO unmatched_raw
        SELECT equipment_raw, 'equipment', 'maintenance', source_file, source_sheet, source_row,
               'Equipment tidak ditemukan di master'
        FROM maintenance_stage
        WHERE equipment_id IS NULL AND nullif(trim(equipment_raw),'') IS NOT NULL
    """)


def _build_rkap_nodes(con: duckdb.DuckDBPyConnection, views: list[str]) -> None:
    if not views:
        return
    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    ru_expr = f"ru_normalize({_cs(c, 'refineryunit','refinery_unit','ru','plant', default='NULL')})"
    con.execute(f"""
        CREATE TABLE rkap_stage AS
        SELECT
            {_cs(c, 'noprogram','no_program','program_no','programkerja','program_kerja','nama_program','program_name')} AS program_no,
            {_cs(c, 'programkerja','program_kerja','nama_program','program_name','noprogram','no_program')} AS program_name,
            {_cs(c, 'equipment','tag_number','tag_no','equnr')} AS equipment_raw,
            {_cs(c, 'katergorirkap','kategori_rkap','kategori')} AS kategori,
            {_cs(c, 'kelompokbiaya','kelompok_biaya')} AS kelompok_biaya,
            {_cs(c, 'disiplin','discipline')} AS disiplin,
            {_cs(c, 'fiscalyear','fiscal_year','tahun', cast=True)} AS fiscal_year,
            {_cs(c, 'planstart','plan_start', cast=True)} AS plan_start,
            {_cs(c, 'planfinish','plan_finish', cast=True)} AS plan_finish,
            {_cs(c, 'actualstart','actual_start', cast=True)} AS actual_start,
            {_cs(c, 'actualfinish','actual_finish', cast=True)} AS actual_finish,
            {_cs(c, 'statusactual','status_actual','status')} AS status_actual,
            {_cs(c, 'totalequivalentidr','total_equivalent_idr', cast=True, default="'0'")} AS total_idr,
            {_cs(c, 'toprisk','top_risk', cast=True)} AS top_risk,
            coalesce({ru_expr}, ru_from_filename(_input_source_file)) AS refinery_unit,
            _input_source_file AS source_file,
            _input_source_sheet AS source_sheet,
            _source_row AS source_row,
            'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
        FROM ({union_sql})
        WHERE nullif(trim({_cs(c, 'noprogram','no_program','program_no','programkerja','program_kerja','nama_program','program_name')}), '') IS NOT NULL
    """)

    con.execute("""
        ALTER TABLE rkap_stage ADD COLUMN program_id VARCHAR;
        UPDATE rkap_stage
        SET program_id = 'node_rkap_program_' || md5(refinery_unit || '|' ||
                          norm_code(program_no) || '|' || norm_code(program_name))
        WHERE refinery_unit IS NOT NULL;

        ALTER TABLE rkap_stage ADD COLUMN equipment_id VARCHAR;
        UPDATE rkap_stage SET equipment_id = e.equipment_id
        FROM equipment_master e
        WHERE rkap_stage.refinery_unit = e.refinery_unit
          AND norm_code(rkap_stage.equipment_raw) = norm_code(e.equipment_code_clean)
          AND nullif(trim(rkap_stage.equipment_raw),'') IS NOT NULL;
    """)

    con.execute("""
        INSERT INTO node_raw
        SELECT DISTINCT program_id, 'rkap_program',
               refinery_unit || '|' || program_no,
               coalesce(nullif(program_name,''), program_no), 'cost_program',
               json_object(
                   'refinery_unit', refinery_unit,
                   'equipment_raw', equipment_raw,
                   'program_no', program_no,
                   'kategori', kategori,
                   'kelompok_biaya', kelompok_biaya,
                   'disiplin', disiplin,
                   'fiscal_year', fiscal_year,
                   'plan_start', plan_start,
                   'plan_finish', plan_finish,
                   'status_actual', status_actual,
                   'derived_total_equivalent_idr_num', total_idr,
                   'derived_is_top_risk', top_risk,
                   'derived_is_delayed',
                       CASE WHEN status_actual ILIKE '%delay%' OR status_actual ILIKE '%terlambat%'
                            THEN 'true' ELSE 'false' END
               ),
               source_file, source_sheet, source_row, source_record_id
        FROM rkap_stage WHERE program_id IS NOT NULL
    """)

    # RU → RKAP
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT r.refinery_unit_id, s.program_id, 'REFINERY_UNIT_HAS_RKAP_PROGRAM',
               'cost_program', 1.0, 'refinery_unit_direct', false,
               json_object('refinery_unit', s.refinery_unit),
               s.source_file, s.source_sheet, s.source_row, s.source_record_id
        FROM rkap_stage s
        JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
        WHERE s.program_id IS NOT NULL
    """)

    # EQUIPMENT → RKAP (exact match)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT equipment_id, program_id, 'EQUIPMENT_HAS_RKAP_PROGRAM',
               'cost_program', 1.0, 'ru_and_equipment_exact', false,
               json_object('match_token', equipment_raw, 'source_ru', refinery_unit),
               source_file, source_sheet, source_row, source_record_id
        FROM rkap_stage WHERE equipment_id IS NOT NULL AND program_id IS NOT NULL
    """)

    # EQUIPMENT → RKAP (token fallback)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT DISTINCT e.equipment_id, s.program_id, 'EQUIPMENT_HAS_RKAP_PROGRAM',
               'cost_program', 0.80, 'rkap_equipment_token_match', true,
               json_object('match_token', trim(token), 'source_ru', s.refinery_unit),
               s.source_file, s.source_sheet, s.source_row, s.source_record_id
        FROM rkap_stage s,
             unnest(string_split(
                 regexp_replace(upper(coalesce(s.equipment_raw,'')), '[,;&]+', ';', 'g'), ';'
             )) t(token)
        JOIN equipment_master e
          ON s.refinery_unit = e.refinery_unit
         AND norm_code(token) = norm_code(e.equipment_code_clean)
        WHERE s.equipment_id IS NULL AND length(norm_code(token)) >= 4
          AND s.program_id IS NOT NULL
    """)


def _build_reliability_nodes(con: duckdb.DuckDBPyConnection, views: list[str]) -> None:
    if not views:
        return
    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    eq_expr = _cs(c, 'equipment','tag_number','tag_no','equnr')
    ru_expr = f"ru_normalize({_cs(c, 'refinery_unit','ru','plant','refineryunit', default='NULL')})"
    con.execute(f"""
        CREATE TABLE reliability_stage AS
        SELECT
            {eq_expr} AS equipment_raw,
            {_cs(c, 'hasil','status','result')} AS hasil,
            {_cs(c, 'running_hours','run_hours','operating_hours','jam_operasi', cast=True, default="'0'")} AS running_hours,
            {_cs(c, 'mtbf', cast=True, default="'0'")} AS mtbf,
            {_cs(c, 'mttr', cast=True, default="'0'")} AS mttr,
            {_cs(c, 'minggu','week', cast=True)} AS minggu,
            {_cs(c, 'bulan','month_update','month','period', cast=True)} AS bulan,
            {_cs(c, 'tahun','year','fiscal_year', cast=True)} AS tahun,
            coalesce({ru_expr}, ru_from_filename(_input_source_file)) AS refinery_unit,
            _input_source_file AS source_file,
            _input_source_sheet AS source_sheet,
            _source_row AS source_row,
            'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
        FROM ({union_sql})
        WHERE nullif(trim({eq_expr}), '') IS NOT NULL
    """)

    con.execute("""
        ALTER TABLE reliability_stage ADD COLUMN obs_id VARCHAR;
        UPDATE reliability_stage
        SET obs_id = 'node_reliability_' || md5(source_record_id);

        ALTER TABLE reliability_stage ADD COLUMN equipment_id VARCHAR;
        UPDATE reliability_stage SET equipment_id = e.equipment_id
        FROM equipment_master e
        WHERE reliability_stage.refinery_unit = e.refinery_unit
          AND norm_code(reliability_stage.equipment_raw) = norm_code(e.equipment_code_clean);
    """)

    con.execute("""
        INSERT INTO node_raw
        SELECT obs_id, 'reliability_observation',
               source_record_id, coalesce(nullif(hasil,''), 'Reliability observation'),
               'reliability',
               json_object(
                   'refinery_unit', refinery_unit,
                   'equipment_raw', equipment_raw,
                   'hasil', hasil,
                   'running_hours', running_hours,
                   'mtbf', mtbf,
                   'mttr', mttr,
                   'minggu', minggu,
                   'bulan', bulan,
                   'tahun', tahun
               ),
               source_file, source_sheet, source_row, source_record_id
        FROM reliability_stage WHERE obs_id IS NOT NULL
    """)

    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT equipment_id, obs_id, 'EQUIPMENT_HAS_RELIABILITY_OBSERVATION',
               'reliability', 1.0, 'ru_and_equipment_exact', false,
               json_object('match_token', equipment_raw),
               source_file, source_sheet, source_row, source_record_id
        FROM reliability_stage WHERE equipment_id IS NOT NULL AND obs_id IS NOT NULL
    """)

    # Node periode (bulan-tahun) + OBSERVED_IN_PERIOD
    con.execute("""
        ALTER TABLE reliability_stage ADD COLUMN period_key VARCHAR;
        UPDATE reliability_stage
        SET period_key = nullif(trim(coalesce(bulan,'') || ' ' || coalesce(tahun,'')), '')
        WHERE bulan IS NOT NULL OR tahun IS NOT NULL;
    """)
    con.execute("""
        INSERT INTO node_raw (node_id, node_type, business_key, label, domain, properties_json)
        SELECT DISTINCT
            'node_period_' || md5(period_key),
            'time_period', period_key, period_key, 'reliability',
            json_object('bulan', bulan, 'tahun', tahun)
        FROM reliability_stage WHERE period_key IS NOT NULL
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT obs_id, 'node_period_' || md5(period_key), 'OBSERVED_IN_PERIOD',
               'reliability', 1.0, 'period_direct', false,
               json_object('period', period_key),
               source_file, source_sheet, source_row, source_record_id
        FROM reliability_stage WHERE obs_id IS NOT NULL AND period_key IS NOT NULL
    """)


def _build_inspection_nodes(con: duckdb.DuckDBPyConnection, views: list[str]) -> None:
    if not views:
        return
    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    eq_expr = _cs(c, 'tag_no_ln','tag_number','tag_no','equipment','equnr')
    ru_expr = f"ru_normalize({_cs(c, 'refinery_unit','ru','plant', default='NULL')})"
    con.execute(f"""
        CREATE TABLE inspection_stage AS
        SELECT
            {eq_expr} AS equipment_raw,
            {_cs(c, 'type_inspection','jenis_inspeksi','inspection_type')} AS type_inspection,
            {_cs(c, 'type_pekerjaan','jenis_pekerjaan','work_type')} AS type_pekerjaan,
            {_cs(c, 'plan_date','tanggal_rencana','inspection_date','next_inspection', cast=True)} AS plan_date,
            {_cs(c, 'actual_date','tanggal_aktual','last_inspection', cast=True)} AS actual_date,
            {_cs(c, 'grand_result','hasil','result','inspection_result')} AS grand_result,
            coalesce({ru_expr}, ru_from_filename(_input_source_file)) AS refinery_unit,
            _input_source_file AS source_file,
            _input_source_sheet AS source_sheet,
            _source_row AS source_row,
            'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
        FROM ({union_sql})
        WHERE nullif(trim({eq_expr}), '') IS NOT NULL
    """)

    con.execute("""
        ALTER TABLE inspection_stage ADD COLUMN inspection_id VARCHAR;
        UPDATE inspection_stage
        SET inspection_id = 'node_inspection_' || md5(source_record_id);

        ALTER TABLE inspection_stage ADD COLUMN equipment_id VARCHAR;
        UPDATE inspection_stage SET equipment_id = e.equipment_id
        FROM equipment_master e
        WHERE inspection_stage.refinery_unit = e.refinery_unit
          AND norm_code(inspection_stage.equipment_raw) = norm_code(e.equipment_code_clean);
    """)

    con.execute("""
        INSERT INTO node_raw
        SELECT inspection_id, 'inspection', source_record_id,
               coalesce(nullif(type_inspection,''), 'Inspection'), 'inspection_issue',
               json_object(
                   'refinery_unit', refinery_unit,
                   'equipment_raw', equipment_raw,
                   'type_inspection', type_inspection,
                   'type_pekerjaan', type_pekerjaan,
                   'plan_date', plan_date,
                   'actual_date', actual_date,
                   'grand_result', grand_result
               ),
               source_file, source_sheet, source_row, source_record_id
        FROM inspection_stage WHERE inspection_id IS NOT NULL
    """)

    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT equipment_id, inspection_id, 'EQUIPMENT_HAS_INSPECTION',
               'inspection_issue', 1.0, 'ru_and_equipment_exact', false,
               json_object('match_token', equipment_raw),
               source_file, source_sheet, source_row, source_record_id
        FROM inspection_stage WHERE equipment_id IS NOT NULL AND inspection_id IS NOT NULL
    """)


def _build_icu_issue_nodes(con: duckdb.DuckDBPyConnection, views: list[str]) -> None:
    if not views:
        return
    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    issue_expr = _cs(c, 'issue', 'deskripsi_issue', 'description')
    ru_expr = f"ru_normalize({_cs(c, 'refinery_unit','ru','kilang','plant', default='NULL')})"
    con.execute(f"""
        CREATE TABLE icu_stage AS
        SELECT
            {_cs(c, 'tag_no','tag_number','equipment','equnr')} AS equipment_raw,
            {issue_expr} AS issue_text,
            {_cs(c, 'icu_status','status')} AS icu_status,
            {_cs(c, 'report_date','tanggal', cast=True)} AS report_date,
            {_cs(c, 'target_closed', cast=True)} AS target_closed,
            {_cs(c, 'mitigation_temporary_solution','mitigation','solusi_sementara')} AS mitigation,
            {_cs(c, 'permanent_solution','solusi_permanen')} AS permanent_solution,
            coalesce(
                {ru_expr},
                ru_from_filename(_input_source_file)
            ) AS refinery_unit,
            _input_source_file AS source_file,
            _input_source_sheet AS source_sheet,
            _source_row AS source_row,
            'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
        FROM ({union_sql})
        WHERE nullif(trim({issue_expr}), '') IS NOT NULL
    """)

    con.execute("""
        ALTER TABLE icu_stage ADD COLUMN issue_id VARCHAR;
        UPDATE icu_stage SET issue_id = 'node_issue_' || md5(source_record_id);

        ALTER TABLE icu_stage ADD COLUMN equipment_id VARCHAR;
        UPDATE icu_stage SET equipment_id = e.equipment_id
        FROM equipment_master e
        WHERE icu_stage.refinery_unit = e.refinery_unit
          AND norm_code(icu_stage.equipment_raw) = norm_code(e.equipment_code_clean)
          AND nullif(trim(icu_stage.equipment_raw),'') IS NOT NULL;
    """)

    con.execute("""
        INSERT INTO node_raw
        SELECT issue_id, 'equipment_issue', source_record_id,
               coalesce(nullif(left(issue_text, 80), ''), 'Issue'), 'inspection_issue',
               json_object(
                   'refinery_unit', refinery_unit,
                   'equipment_raw', equipment_raw,
                   'issue', issue_text,
                   'icu_status', icu_status,
                   'report_date', report_date,
                   'target_closed', target_closed,
                   'mitigation', mitigation,
                   'permanent_solution', permanent_solution
               ),
               source_file, source_sheet, source_row, source_record_id
        FROM icu_stage WHERE issue_id IS NOT NULL
    """)

    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT equipment_id, issue_id, 'EQUIPMENT_HAS_ISSUE',
               'inspection_issue', 1.0, 'ru_and_equipment_exact', false,
               json_object('match_token', equipment_raw),
               source_file, source_sheet, source_row, source_record_id
        FROM icu_stage WHERE equipment_id IS NOT NULL AND issue_id IS NOT NULL
    """)

    # RU → Issue shortcut
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT r.refinery_unit_id, s.issue_id, 'REFINERY_UNIT_HAS_ISSUE',
               'inspection_issue', 1.0, 'refinery_unit_direct', false,
               json_object('refinery_unit', s.refinery_unit),
               s.source_file, s.source_sheet, s.source_row, s.source_record_id
        FROM icu_stage s
        JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
        WHERE s.issue_id IS NOT NULL
    """)


def _build_readiness_nodes(con: duckdb.DuckDBPyConnection, views: list[str]) -> None:
    if not views:
        return
    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    ru_expr = f"ru_normalize({_cs(c, 'refinery_unit','ru','kilang','plant', default='NULL')})"
    con.execute(f"""
        CREATE TABLE readiness_stage AS
        SELECT
            {_cs(c, 'equipment','tag_number','tag_no','process_equipment','nama_tangki','no_tangki')} AS equipment_raw,
            {_cs(c, 'period_date','month_update', cast=True)} AS period_date,
            {_cs(c, 'status_operation','status_operasi')} AS status_operation,
            {_cs(c, 'status_item')} AS status_item,
            {_cs(c, 'remark','keterangan')} AS remark,
            {_cs(c, 'rtl')} AS rtl,
            coalesce(
                {ru_expr},
                ru_from_filename(_input_source_file)
            ) AS refinery_unit,
            _input_source_file AS source_file,
            _input_source_sheet AS source_sheet,
            _source_row AS source_row,
            'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
        FROM ({union_sql})
    """)

    con.execute("""
        ALTER TABLE readiness_stage ADD COLUMN readiness_id VARCHAR;
        UPDATE readiness_stage SET readiness_id = 'node_readiness_' || md5(source_record_id);

        ALTER TABLE readiness_stage ADD COLUMN equipment_id VARCHAR;
        UPDATE readiness_stage SET equipment_id = e.equipment_id
        FROM equipment_master e
        WHERE readiness_stage.refinery_unit = e.refinery_unit
          AND norm_code(readiness_stage.equipment_raw) = norm_code(e.equipment_code_clean)
          AND nullif(trim(readiness_stage.equipment_raw),'') IS NOT NULL;
    """)

    con.execute("""
        INSERT INTO node_raw
        SELECT readiness_id, 'readiness_record', source_record_id,
               coalesce(nullif(status_operation,''), 'Readiness record'), 'readiness_operation',
               json_object(
                   'refinery_unit', refinery_unit,
                   'equipment_raw', equipment_raw,
                   'period_date', period_date,
                   'status_operation', status_operation,
                   'status_item', status_item,
                   'remark', remark,
                   'rtl', rtl
               ),
               source_file, source_sheet, source_row, source_record_id
        FROM readiness_stage WHERE readiness_id IS NOT NULL
    """)

    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT equipment_id, readiness_id, 'EQUIPMENT_HAS_READINESS_RECORD',
               'readiness_operation', 1.0, 'ru_and_equipment_exact', false,
               json_object('match_token', equipment_raw),
               source_file, source_sheet, source_row, source_record_id
        FROM readiness_stage WHERE equipment_id IS NOT NULL AND readiness_id IS NOT NULL
    """)

    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT r.refinery_unit_id, s.readiness_id, 'REFINERY_UNIT_HAS_READINESS_RECORD',
               'readiness_operation', 1.0, 'refinery_unit_direct', false,
               json_object('refinery_unit', s.refinery_unit),
               s.source_file, s.source_sheet, s.source_row, s.source_record_id
        FROM readiness_stage s
        JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
        WHERE s.readiness_id IS NOT NULL
    """)


# ---------------------------------------------------------------------------
# OA Data: Operational Availability, Allowance Unplanned, Issue List
# ---------------------------------------------------------------------------

def _build_oa_nodes(con: duckdb.DuckDBPyConnection,
                    allowance_views: list[str],
                    availability_views: list[str],
                    issue_views: list[str]) -> None:
    con.execute("""
        CREATE TABLE IF NOT EXISTS oa_allowance_stage (
            refinery_unit VARCHAR, mi_pi VARCHAR, allowance_day DOUBLE, unplanned_day DOUBLE, month_update VARCHAR
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS oa_availability_stage (
            refinery_unit VARCHAR, actual_target VARCHAR, value_perc DOUBLE, month_update VARCHAR
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS oa_issue_stage (
            refinery_unit VARCHAR, mi_pi VARCHAR, issue VARCHAR, month_update VARCHAR, date_issue VARCHAR
        )
    """)

    for tname in allowance_views:
        cols = [r[0].lower() for r in con.execute(f"DESCRIBE {tname}").fetchall()]
        ru_col  = next((c for c in cols if 'refinery' in c), cols[0])
        mip_col = next((c for c in cols if 'mi' in c or 'pi' in c), cols[1] if len(cols) > 1 else cols[0])
        alw_col = next((c for c in cols if 'allowance' in c), cols[2] if len(cols) > 2 else cols[0])
        unp_col = next((c for c in cols if 'unplanned' in c), cols[3] if len(cols) > 3 else cols[0])
        mon_col = next((c for c in cols if 'month' in c or 'update' in c), cols[4] if len(cols) > 4 else cols[0])
        con.execute(f"""
            INSERT INTO oa_allowance_stage
            SELECT TRIM(CAST("{ru_col}" AS VARCHAR)), TRIM(CAST("{mip_col}" AS VARCHAR)),
                   TRY_CAST("{alw_col}" AS DOUBLE), TRY_CAST("{unp_col}" AS DOUBLE),
                   CAST("{mon_col}" AS VARCHAR)
            FROM {tname} WHERE "{ru_col}" IS NOT NULL
        """)

    for tname in availability_views:
        cols = [r[0].lower() for r in con.execute(f"DESCRIBE {tname}").fetchall()]
        ru_col  = next((c for c in cols if 'refinery' in c), cols[0])
        at_col  = next((c for c in cols if 'actual' in c or 'target' in c), cols[1] if len(cols) > 1 else cols[0])
        val_col = next((c for c in cols if 'value' in c or 'perc' in c), cols[2] if len(cols) > 2 else cols[0])
        mon_col = next((c for c in cols if 'month' in c or 'update' in c), cols[3] if len(cols) > 3 else cols[0])
        con.execute(f"""
            INSERT INTO oa_availability_stage
            SELECT TRIM(CAST("{ru_col}" AS VARCHAR)), TRIM(CAST("{at_col}" AS VARCHAR)),
                   TRY_CAST("{val_col}" AS DOUBLE), CAST("{mon_col}" AS VARCHAR)
            FROM {tname} WHERE "{ru_col}" IS NOT NULL
        """)

    for tname in issue_views:
        cols = [r[0].lower() for r in con.execute(f"DESCRIBE {tname}").fetchall()]
        ru_col  = next((c for c in cols if 'refinery' in c), cols[0])
        mip_col = next((c for c in cols if 'mi' in c or 'pi' in c), cols[1] if len(cols) > 1 else cols[0])
        iss_col = next((c for c in cols if 'issue' in c), cols[2] if len(cols) > 2 else cols[0])
        mon_col = next((c for c in cols if 'month' in c), cols[3] if len(cols) > 3 else cols[0])
        dat_col = next((c for c in cols if 'date' in c), cols[4] if len(cols) > 4 else cols[0])
        con.execute(f"""
            INSERT INTO oa_issue_stage
            SELECT TRIM(CAST("{ru_col}" AS VARCHAR)), TRIM(CAST("{mip_col}" AS VARCHAR)),
                   TRIM(CAST("{iss_col}" AS VARCHAR)), CAST("{mon_col}" AS VARCHAR),
                   CAST("{dat_col}" AS VARCHAR)
            FROM {tname} WHERE "{ru_col}" IS NOT NULL AND "{iss_col}" IS NOT NULL
        """)

    # OA availability nodes
    con.execute("""
        INSERT INTO node_raw (node_id, node_type, business_key, label, domain,
                              properties_json, source_file, source_sheet, source_row, source_record_id)
        SELECT
            'oa_avail_' || norm_code(a.refinery_unit || '_' || a.actual_target || '_' || a.month_update),
            'oa_availability', a.refinery_unit || '|' || a.actual_target,
            a.refinery_unit || ' OA ' || a.actual_target || ' (' || LEFT(a.month_update, 7) || ')',
            'oa',
            json_object('refinery_unit', a.refinery_unit, 'actual_target', a.actual_target,
                        'value_perc', COALESCE(CAST(a.value_perc AS VARCHAR), ''),
                        'month_update', a.month_update),
            'oa_data', 'Operational Availability', row_number() OVER (), ''
        FROM oa_availability_stage a WHERE a.refinery_unit IS NOT NULL
    """)

    # OA issue nodes
    con.execute("""
        INSERT INTO node_raw (node_id, node_type, business_key, label, domain,
                              properties_json, source_file, source_sheet, source_row, source_record_id)
        SELECT
            'oa_issue_' || norm_code(oi.refinery_unit || '_' || oi.issue || '_' || oi.month_update),
            'oa_issue', oi.refinery_unit || '|' || LEFT(oi.issue, 40),
            LEFT(oi.issue, 80), 'oa',
            json_object('refinery_unit', oi.refinery_unit, 'mi_pi', oi.mi_pi,
                        'issue', oi.issue, 'month_update', oi.month_update,
                        'date_issue', oi.date_issue),
            'oa_data', 'Issue List', row_number() OVER (), ''
        FROM oa_issue_stage oi
    """)

    # Edges: RU → OA availability
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT DISTINCT r.refinery_unit_id, n.node_id, 'REFINERY_UNIT_HAS_OA_AVAILABILITY',
               'oa', 1.0, 'refinery_unit_direct', false,
               json_object('refinery_unit', r.refinery_unit),
               'oa_data', 'Operational Availability', 0, ''
        FROM node_raw n
        JOIN oa_availability_stage a
          ON n.node_type = 'oa_availability'
         AND json_extract_string(n.properties_json, '$.refinery_unit') = a.refinery_unit
        JOIN ru_reference r ON norm_text(a.refinery_unit) = norm_text(r.refinery_unit)
    """)

    # Edges: RU → OA issue
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT DISTINCT r.refinery_unit_id, n.node_id, 'REFINERY_UNIT_HAS_OA_ISSUE',
               'oa', 1.0, 'refinery_unit_direct', false,
               json_object('refinery_unit', r.refinery_unit),
               'oa_data', 'Issue List', 0, ''
        FROM node_raw n
        JOIN oa_issue_stage oi
          ON n.node_type = 'oa_issue'
         AND json_extract_string(n.properties_json, '$.refinery_unit') = oi.refinery_unit
         AND json_extract_string(n.properties_json, '$.issue') = oi.issue
        JOIN ru_reference r ON norm_text(oi.refinery_unit) = norm_text(r.refinery_unit)
    """)


# ---------------------------------------------------------------------------
# PLO: Perizinan Layak Operasi
# ---------------------------------------------------------------------------

def _build_plo_nodes(con: duckdb.DuckDBPyConnection, views: list[str]) -> None:
    if not views:
        return

    con.execute("""
        CREATE TABLE IF NOT EXISTS plo_stage (
            refinery_unit VARCHAR, nomor_ijin VARCHAR, nama_plo VARCHAR,
            kapasitas VARCHAR, date_expired VARCHAR, days_expired INTEGER, status_plo VARCHAR
        )
    """)

    for tname in views:
        cols = [r[0].lower() for r in con.execute(f"DESCRIBE {tname}").fetchall()]
        ru_col   = next((c for c in cols if 'refinery' in c), cols[0])
        no_col   = next((c for c in cols if 'nomor' in c or 'ijin' in c or 'number' in c), cols[1] if len(cols) > 1 else cols[0])
        nm_col   = next((c for c in cols if 'nama' in c or 'name' in c), cols[2] if len(cols) > 2 else cols[0])
        kap_col  = next((c for c in cols if 'cakupan' in c or 'kapasitas' in c or 'capacity' in c), cols[3] if len(cols) > 3 else cols[0])
        exp_col  = next((c for c in cols if 'expired' in c and 'sum' not in c and 'days' not in c and 'day' not in c), cols[4] if len(cols) > 4 else cols[0])
        days_col = next((c for c in cols if 'days' in c or ('sum' in c and 'expired' in c)), cols[5] if len(cols) > 5 else cols[0])
        st_col   = next((c for c in cols if 'status' in c), cols[6] if len(cols) > 6 else cols[0])
        con.execute(f"""
            INSERT INTO plo_stage
            SELECT TRIM(CAST("{ru_col}" AS VARCHAR)), TRIM(CAST("{no_col}" AS VARCHAR)),
                   TRIM(CAST("{nm_col}" AS VARCHAR)), TRIM(CAST("{kap_col}" AS VARCHAR)),
                   CAST("{exp_col}" AS VARCHAR), TRY_CAST("{days_col}" AS INTEGER),
                   TRIM(CAST("{st_col}" AS VARCHAR))
            FROM {tname}
            WHERE "{ru_col}" ILIKE '%RU%' AND "{no_col}" IS NOT NULL AND "{no_col}" != ''
        """)

    con.execute("""
        INSERT INTO node_raw (node_id, node_type, business_key, label, domain,
                              properties_json, source_file, source_sheet, source_row, source_record_id)
        SELECT
            'plo_' || norm_code(p.nomor_ijin), 'plo_permit', p.nomor_ijin,
            LEFT(p.nama_plo, 80), 'plo',
            json_object('refinery_unit', p.refinery_unit, 'nomor_ijin', p.nomor_ijin,
                        'nama_plo', p.nama_plo, 'kapasitas', p.kapasitas,
                        'date_expired', p.date_expired,
                        'days_expired', COALESCE(CAST(p.days_expired AS VARCHAR), ''),
                        'status_plo', COALESCE(p.status_plo, '')),
            'plo', 'Sheet1', row_number() OVER (), p.nomor_ijin
        FROM plo_stage p WHERE p.nomor_ijin IS NOT NULL AND p.nomor_ijin != ''
    """)

    # Edges: RU → PLO permit (kedua sisi dinormalisasi ru_normalize → 'RU II Dumai')
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT DISTINCT r.refinery_unit_id, n.node_id, 'REFINERY_UNIT_HAS_PLO_PERMIT',
               'plo', 1.0, 'refinery_unit_direct', false,
               json_object('refinery_unit', r.refinery_unit, 'status_plo', p.status_plo),
               'plo', 'Sheet1', 0, ''
        FROM node_raw n
        JOIN plo_stage p ON n.business_key = p.nomor_ijin AND n.node_type = 'plo_permit'
        JOIN ru_reference r
          ON ru_normalize(p.refinery_unit) = r.refinery_unit
    """)


# ---------------------------------------------------------------------------
# RCPS: Root Cause Problem Solving + rekomendasi
# ---------------------------------------------------------------------------

def _build_rcps_nodes(con: duckdb.DuckDBPyConnection,
                      rcps_views: list[str], rec_views: list[str]) -> None:
    # File RCPS Pertamina berformat gabungan: header RCPS + rekomendasi dalam
    # satu tabel (tiap baris = 1 rekomendasi, dikelompokkan per RCPS No). Kedua
    # domain (rcps / rcps_recommendation) diperlakukan sebagai sumber baris yang
    # sama karena strukturnya identik.
    views = list(dict.fromkeys(rcps_views + rec_views))
    if not views:
        return

    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    rcps_no = _cs(c, 'rcps_no','norcps','no_rcps','rcps_number','id_rcps')
    ru_expr = f"ru_normalize({_cs(c, 'kilang','refinery_unit','ru','plant','refineryunit', default='NULL')})"
    con.execute(f"""
        CREATE TABLE rcps_stage AS
        SELECT
            {rcps_no} AS rcps_no,
            {_cs(c, 'judul_rcps','judul','title','problem','masalah')} AS judul_rcps,
            {_cs(c, 'link_rcps','link','url')} AS link_rcps,
            {_cs(c, 'description','deskripsi','recommendation','rekomendasi','saran','tindak_lanjut','action')} AS rec_text,
            {_cs(c, 'recomendation','recommendation_category','kategori_rekomendasi','recommendation_type','recommendationcategory')} AS rec_category,
            {_cs(c, 'traffic','traffic_light','warna')} AS traffic,
            {_cs(c, 'pic','responsible','penanggung_jawab')} AS pic,
            {_cs(c, 'target','target_date','due_date','tanggal_target')} AS target_date,
            {_cs(c, 'no_irkap','noirkap','irkap_no','no_rkap','irkap')} AS no_irkap,
            {_cs(c, 'remark','keterangan','status','catatan')} AS remark,
            coalesce({ru_expr}, ru_from_filename(_input_source_file)) AS refinery_unit,
            _input_source_file AS source_file,
            _input_source_sheet AS source_sheet,
            _source_row AS source_row,
            'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
        FROM ({union_sql})
        WHERE nullif(trim({rcps_no}), '') IS NOT NULL
    """)
    # rcps_id dikelompokkan per (RU + nomor RCPS) — satu node RCPS per kasus
    con.execute("""
        ALTER TABLE rcps_stage ADD COLUMN rcps_id VARCHAR;
        UPDATE rcps_stage
        SET rcps_id = 'node_rcps_' || md5(coalesce(refinery_unit,'UNKNOWN') || '|' || norm_code(rcps_no));

        ALTER TABLE rcps_stage ADD COLUMN rec_id VARCHAR;
        UPDATE rcps_stage SET rec_id = 'node_recommendation_' || md5(source_record_id);
    """)

    # Node RCPS (unik per nomor RCPS, judul dari Judul RCPS)
    con.execute("""
        INSERT INTO node_raw
        SELECT DISTINCT ON (rcps_id)
               rcps_id, 'rcps', rcps_no,
               coalesce(nullif(left(judul_rcps,90),''), rcps_no), 'rcps',
               json_object('refinery_unit', refinery_unit, 'rcps_no', rcps_no,
                   'judul_rcps', judul_rcps, 'link_rcps', link_rcps),
               source_file, source_sheet, source_row, source_record_id
        FROM rcps_stage WHERE rcps_id IS NOT NULL
    """)

    # Node rekomendasi (per baris)
    con.execute("""
        INSERT INTO node_raw
        SELECT rec_id, 'recommendation', source_record_id,
               coalesce(nullif(left(rec_text,90),''), 'Rekomendasi'), 'rcps',
               json_object('rcps_no', rcps_no, 'recommendation', rec_text,
                   'rec_category', rec_category, 'traffic', traffic, 'pic', pic,
                   'target_date', target_date, 'no_irkap', no_irkap, 'remark', remark),
               source_file, source_sheet, source_row, source_record_id
        FROM rcps_stage WHERE rec_id IS NOT NULL AND nullif(trim(rec_text),'') IS NOT NULL
    """)

    # RCPS_HAS_RECOMMENDATION
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT rcps_id, rec_id, 'RCPS_HAS_RECOMMENDATION',
               'rcps', 1.0, 'rcps_no_group', false,
               json_object('rcps_no', rcps_no),
               source_file, source_sheet, source_row, source_record_id
        FROM rcps_stage
        WHERE rcps_id IS NOT NULL AND rec_id IS NOT NULL
          AND nullif(trim(rec_text),'') IS NOT NULL
    """)

    # REFINERY_UNIT_HAS_RCPS
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT DISTINCT r.refinery_unit_id, s.rcps_id, 'REFINERY_UNIT_HAS_RCPS',
               'rcps', 1.0, 'refinery_unit_direct', false,
               json_object('refinery_unit', s.refinery_unit),
               s.source_file, s.source_sheet, s.source_row, s.source_record_id
        FROM rcps_stage s
        JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
        WHERE s.rcps_id IS NOT NULL
    """)

    # RCPS → RKAP program via No. IRKAP (jika ada)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT DISTINCT s.rcps_id, n.node_id, 'RCPS_HAS_RKAP_PROGRAM',
               'rcps', 0.9, 'irkap_no_match', false,
               json_object('no_irkap', s.no_irkap),
               s.source_file, s.source_sheet, s.source_row, s.source_record_id
        FROM rcps_stage s
        JOIN node_raw n
          ON n.node_type = 'rkap_program'
         AND norm_code(s.no_irkap) = norm_code(n.properties_json ->> 'program_no')
         AND nullif(trim(s.no_irkap), '') IS NOT NULL
        WHERE s.rcps_id IS NOT NULL
    """)


# ---------------------------------------------------------------------------
# Node builders — domain tambahan
# ---------------------------------------------------------------------------


def _build_work_order_nodes(con: duckdb.DuckDBPyConnection, views: list[str]) -> None:
    if not views:
        return
    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    ord_expr = _cs(c, 'order_no', 'aufnr', 'order', 'order_number', 'maint_order', cast=True, default='NULL')
    notif_expr = _cs(c, 'notification_no', 'notification', 'notif', 'qmnum', cast=True, default='NULL')
    ru_expr = f"ru_normalize({_cs(c, 'refinery_unit','ru','plant','maintplant','location', default='NULL')})"
    eq_raw_wo = _cs(c, 'equipment', default='NULL')
    eq_clean_wo = f"regexp_replace(trim({eq_raw_wo}), '/[0-9]+$', '')" if eq_raw_wo != 'NULL' else 'NULL'
    con.execute(f"""
        CREATE TABLE work_order_stage AS
        SELECT
            {ord_expr} AS order_raw,
            {notif_expr} AS notification_raw,
            coalesce(norm_code({ord_expr}), norm_code({notif_expr})) AS order_code,
            {_cs(c, 'description','kurztext','short_text','order_description')} AS order_desc,
            {_cs(c, 'order_type','auart','order_category')} AS order_type,
            {_cs(c, 'priority','priok')} AS priority,
            {_cs(c, 'user_status','txt04','ustatus')} AS user_status,
            {_cs(c, 'system_status','sttxt')} AS system_status,
            {_cs(c, 'basic_start','actual_start','gstrp','req_start','basic_start_date', cast=True)} AS reference_date,
            {_cs(c, 'total_planned_costs','geplk','planned_cost', cast=True, default="'0'")} AS planned_cost,
            {_cs(c, 'total_actual_costs','istko','actual_cost', cast=True, default="'0'")} AS actual_cost,
            {_cs(c, 'main_work_center','main_workcenter','arbpl','work_center','workcenter','main_workctr')} AS work_center,
            {eq_raw_wo} AS equipment_raw,
            {eq_clean_wo} AS equipment_clean,
            coalesce({ru_expr}, ru_from_filename(_input_source_file)) AS refinery_unit,
            _input_source_file AS source_file,
            _input_source_sheet AS source_sheet,
            _source_row AS source_row,
            'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
        FROM ({union_sql})
        WHERE {ord_expr} IS NOT NULL
    """)
    con.execute("""
        ALTER TABLE work_order_stage ADD COLUMN wo_id VARCHAR;
        UPDATE work_order_stage
        SET wo_id = 'node_wo_' || md5(coalesce(refinery_unit, 'UNKNOWN') || '|' || order_code)
        WHERE order_code IS NOT NULL;

        ALTER TABLE work_order_stage ADD COLUMN equipment_id VARCHAR;
    """)
    has_eq = con.execute("SELECT count(*) FROM information_schema.tables WHERE table_name='equipment_master'").fetchone()[0]
    if has_eq:
        con.execute("""
            UPDATE work_order_stage SET equipment_id = e.equipment_id
            FROM equipment_master e
            WHERE work_order_stage.equipment_clean = e.equipment_code_clean
              AND work_order_stage.equipment_clean IS NOT NULL;
        """)
    con.execute("""
        INSERT INTO node_raw
        SELECT DISTINCT wo_id, 'work_order', refinery_unit || '|' || order_code,
               coalesce(nullif(order_desc,''), order_raw), 'maintenance',
               json_object(
                   'refinery_unit', refinery_unit,
                   'equipment_raw', equipment_raw,
                   'order_raw', order_raw,
                   'notification_raw', notification_raw,
                   'order_type', order_type,
                   'priority', priority,
                   'user_status', user_status,
                   'system_status', system_status,
                   'reference_date', reference_date,
                   'derived_planned_cost', planned_cost,
                   'derived_actual_cost', actual_cost,
                   'work_center', work_center,
                   'derived_is_open_order',
                       CASE WHEN user_status ILIKE '%WAMA%' OR user_status ILIKE '%WASR%'
                                 OR system_status ILIKE '%REL%' OR system_status ILIKE '%PCNF%'
                            THEN 'true' ELSE 'false' END,
                   'derived_status_bucket',
                       CASE WHEN user_status ILIKE '%WAMA%' THEN 'WAMA'
                            WHEN user_status ILIKE '%WASR%' THEN 'WASR'
                            WHEN system_status ILIKE '%TECO%' THEN 'TECO'
                            WHEN system_status ILIKE '%CLSD%' THEN 'CLSD'
                            ELSE 'OPEN' END
               ),
               source_file, source_sheet, source_row, source_record_id
        FROM work_order_stage WHERE wo_id IS NOT NULL
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT equipment_id, wo_id, 'EQUIPMENT_HAS_WORK_ORDER',
               'maintenance', 1.0, 'ru_and_equipment_exact', false,
               json_object('match_token', equipment_raw, 'source_ru', refinery_unit),
               source_file, source_sheet, source_row, source_record_id
        FROM work_order_stage WHERE equipment_id IS NOT NULL AND wo_id IS NOT NULL
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT r.refinery_unit_id, s.wo_id, 'REFINERY_UNIT_HAS_WORK_ORDER',
               'maintenance', 1.0, 'refinery_unit_direct', false,
               json_object('refinery_unit', s.refinery_unit),
               s.source_file, s.source_sheet, s.source_row, s.source_record_id
        FROM work_order_stage s
        JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
        WHERE s.wo_id IS NOT NULL
    """)
    con.execute("""
        INSERT INTO unmatched_raw
        SELECT equipment_raw, 'equipment', 'work_order', source_file, source_sheet, source_row,
               'Equipment tidak ditemukan di master'
        FROM work_order_stage
        WHERE equipment_id IS NULL AND nullif(trim(equipment_raw),'') IS NOT NULL
    """)


def _build_notification_nodes(con: duckdb.DuckDBPyConnection, views: list[str]) -> None:
    if not views:
        return
    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    notif_expr = _cs(c, 'notification_no', 'qmnum', 'notif_no', 'notification', 'notif', cast=True, default='NULL')
    ru_expr = f"ru_normalize({_cs(c, 'refinery_unit','ru','plant','maintplant','location', default='NULL')})"
    # Equipment SAP notif sering punya suffix /00 — strip sebelum norm_code
    eq_raw = _cs(c, 'equipment', default='NULL')
    eq_clean = f"regexp_replace(trim({eq_raw}), '/[0-9]+$', '')" if eq_raw != 'NULL' else 'NULL'
    con.execute(f"""
        CREATE TABLE notification_stage AS
        SELECT
            {notif_expr} AS notif_raw,
            norm_code({notif_expr}) AS notif_code,
            {_cs(c, 'short_text','description','kurztext','notif_description')} AS notif_desc,
            {_cs(c, 'notifictn_type','notif_type','notification_type','qmart')} AS notif_type,
            {_cs(c, 'priority','priok')} AS priority,
            {_cs(c, 'system_status','sttxt')} AS system_status,
            {_cs(c, 'user_status','txt04','ustatus')} AS user_status,
            {_cs(c, 'notification_date','malfunction_start','reported_on', cast=True)} AS notif_date,
            {_cs(c, 'malfunction_end','end_date', cast=True)} AS end_date,
            {_cs(c, 'breakdown','breakdown_indicator')} AS breakdown,
            {eq_raw} AS equipment_raw,
            {eq_clean} AS equipment_clean,
            {_cs(c, 'functional_loc','functional_location','floc')} AS functional_loc,
            coalesce({ru_expr}, ru_from_filename(_input_source_file)) AS refinery_unit,
            _input_source_file AS source_file,
            _input_source_sheet AS source_sheet,
            _source_row AS source_row,
            'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
        FROM ({union_sql})
        WHERE {notif_expr} IS NOT NULL
    """)
    con.execute("""
        ALTER TABLE notification_stage ADD COLUMN notif_id VARCHAR;
        UPDATE notification_stage
        SET notif_id = 'node_notif_' || md5(coalesce(refinery_unit, 'UNKNOWN') || '|' || notif_code)
        WHERE notif_code IS NOT NULL;

        ALTER TABLE notification_stage ADD COLUMN equipment_id VARCHAR;
    """)
    has_eq = con.execute("SELECT count(*) FROM information_schema.tables WHERE table_name='equipment_master'").fetchone()[0]
    if has_eq:
        con.execute("""
            UPDATE notification_stage SET equipment_id = e.equipment_id
            FROM equipment_master e
            WHERE notification_stage.equipment_clean = e.equipment_code_clean
              AND notification_stage.equipment_clean IS NOT NULL;
        """)
    con.execute("""
        INSERT INTO node_raw
        SELECT DISTINCT notif_id, 'notification', refinery_unit || '|' || notif_code,
               coalesce(nullif(notif_desc,''), notif_raw), 'maintenance',
               json_object(
                   'refinery_unit', refinery_unit,
                   'equipment_raw', equipment_raw,
                   'notif_raw', notif_raw,
                   'notif_type', notif_type,
                   'priority', priority,
                   'system_status', system_status,
                   'user_status', user_status,
                   'notif_date', notif_date,
                   'end_date', end_date,
                   'breakdown', breakdown,
                   'functional_loc', functional_loc
               ),
               source_file, source_sheet, source_row, source_record_id
        FROM notification_stage WHERE notif_id IS NOT NULL
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT equipment_id, notif_id, 'EQUIPMENT_HAS_NOTIFICATION',
               'maintenance', 1.0, 'ru_and_equipment_exact', false,
               json_object('match_token', equipment_raw, 'source_ru', refinery_unit),
               source_file, source_sheet, source_row, source_record_id
        FROM notification_stage WHERE equipment_id IS NOT NULL AND notif_id IS NOT NULL
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT r.refinery_unit_id, s.notif_id, 'REFINERY_UNIT_HAS_NOTIFICATION',
               'maintenance', 1.0, 'refinery_unit_direct', false,
               json_object('refinery_unit', s.refinery_unit),
               s.source_file, s.source_sheet, s.source_row, s.source_record_id
        FROM notification_stage s
        JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
        WHERE s.notif_id IS NOT NULL
    """)
    # Hubungkan Work Order → Notification jika ada tabel work_order_stage
    try:
        con.execute("""
            INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
                domain, confidence, match_rule, is_candidate, properties_json,
                source_file, source_sheet, source_row, source_record_id)
            SELECT w.wo_id, n.notif_id, 'WORK_ORDER_HAS_NOTIFICATION',
                   'maintenance', 1.0, 'wo_notif_link', false,
                   json_object('notification', n.notif_raw),
                   n.source_file, n.source_sheet, n.source_row, n.source_record_id
            FROM notification_stage n
            JOIN work_order_stage w
              ON norm_code(n.notif_raw) = norm_code(w.notification_raw)
             AND n.refinery_unit = w.refinery_unit
            WHERE n.notif_id IS NOT NULL AND w.wo_id IS NOT NULL
        """)
    except Exception:
        pass
    con.execute("""
        INSERT INTO unmatched_raw
        SELECT equipment_raw, 'equipment', 'notification', source_file, source_sheet, source_row,
               'Equipment tidak ditemukan di master'
        FROM notification_stage
        WHERE equipment_id IS NULL AND nullif(trim(equipment_raw),'') IS NOT NULL
    """)


def _build_workplan_nodes(
    con: duckdb.DuckDBPyConnection,
    views: list[str],
    domain_key: str,           # 'spm_workplan' / 'tank_workplan' / 'jetty_workplan'
    node_type: str,            # 'spm_workplan' etc.
    domain_label: str,         # 'SPM' / 'Tank' / 'Jetty'
    stage_table: str,          # 'spm_workplan_stage' etc.
    rel_type: str,             # 'EQUIPMENT_HAS_SPM_WORKPLAN' etc.
    ru_rel_type: str,          # 'REFINERY_UNIT_HAS_SPM_WORKPLAN' etc.
) -> None:
    if not views:
        return
    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    ru_expr = f"ru_normalize({_cs(c, 'refinery_unit','ru','plant', default='NULL')})"
    name_expr = _cs(c, 'nama_program','program_name','program','nama_kegiatan','kegiatan', default='NULL')
    period_expr = _cs(c, 'period','tahun','year','bulan','month','tanggal','date', cast=True, default='NULL')
    con.execute(f"""
        CREATE TABLE {stage_table} AS
        SELECT
            {_cs(c, 'equipment')} AS equipment_raw,
            {name_expr} AS program_name,
            {period_expr} AS period_date,
            {_cs(c, 'target','target_realisasi','target_progres')} AS target,
            {_cs(c, 'realisasi','progres','actual','pencapaian')} AS realisasi,
            {_cs(c, 'status','status_program','keterangan')} AS status,
            {_cs(c, 'remark','catatan','keterangan_tambahan')} AS remark,
            coalesce({ru_expr}, ru_from_filename(_input_source_file)) AS refinery_unit,
            _input_source_file AS source_file,
            _input_source_sheet AS source_sheet,
            _source_row AS source_row,
            'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
        FROM ({union_sql})
    """)
    con.execute(f"""
        ALTER TABLE {stage_table} ADD COLUMN wp_id VARCHAR;
        UPDATE {stage_table}
        SET wp_id = 'node_{node_type}_' || md5(source_record_id);

        ALTER TABLE {stage_table} ADD COLUMN equipment_id VARCHAR;
        UPDATE {stage_table} SET equipment_id = e.equipment_id
        FROM equipment_master e
        WHERE {stage_table}.refinery_unit = e.refinery_unit
          AND norm_code({stage_table}.equipment_raw) = norm_code(e.equipment_code_clean)
          AND {stage_table}.equipment_raw IS NOT NULL;
    """)
    con.execute(f"""
        INSERT INTO node_raw
        SELECT wp_id, '{node_type}', source_record_id,
               coalesce(nullif(program_name,''), '{domain_label} Workplan'), '{node_type}',
               json_object(
                   'refinery_unit', refinery_unit,
                   'equipment_raw', equipment_raw,
                   'program_name', program_name,
                   'period_date', period_date,
                   'target', target,
                   'realisasi', realisasi,
                   'status', status,
                   'remark', remark
               ),
               source_file, source_sheet, source_row, source_record_id
        FROM {stage_table} WHERE wp_id IS NOT NULL
    """)
    con.execute(f"""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT equipment_id, wp_id, '{rel_type}',
               '{node_type}', 1.0, 'ru_and_equipment_exact', false,
               json_object('match_token', equipment_raw),
               source_file, source_sheet, source_row, source_record_id
        FROM {stage_table} WHERE equipment_id IS NOT NULL AND wp_id IS NOT NULL
    """)
    con.execute(f"""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT r.refinery_unit_id, s.wp_id, '{ru_rel_type}',
               '{node_type}', 1.0, 'refinery_unit_direct', false,
               json_object('refinery_unit', s.refinery_unit),
               s.source_file, s.source_sheet, s.source_row, s.source_record_id
        FROM {stage_table} s
        JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
        WHERE s.wp_id IS NOT NULL
    """)


def _build_readiness_subtype_nodes(
    con: duckdb.DuckDBPyConnection,
    views: list[str],
    node_type: str,           # 'readiness_tank' etc.
    domain_label: str,        # 'Tank' etc.
    stage_table: str,         # 'readiness_tank_stage' etc.
    extra_cols: list[str],    # extra columns to include in properties_json
) -> None:
    if not views:
        return
    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    ru_expr = f"ru_normalize({_cs(c, 'refinery_unit','ru','plant', default='NULL')})"
    # Build extra properties dynamically
    extra_props = ""
    for col in extra_cols:
        if col in c:
            extra_props += f", '{col}', {col}"
    con.execute(f"""
        CREATE TABLE {stage_table} AS
        SELECT
            {_cs(c, 'equipment','tag_number','tag_no','nama_tangki','no_tangki','nama_spm','nama_dermaga')} AS equipment_raw,
            {_cs(c, 'period_date','month_update','bulan','tanggal', cast=True)} AS period_date,
            {_cs(c, 'status_operation','status_operasi')} AS status_operation,
            {_cs(c, 'status_item')} AS status_item,
            {_cs(c, 'remark','keterangan','catatan')} AS remark,
            {_cs(c, 'rtl')} AS rtl,
            coalesce({ru_expr}, ru_from_filename(_input_source_file)) AS refinery_unit,
            _input_source_file AS source_file,
            _input_source_sheet AS source_sheet,
            _source_row AS source_row,
            'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
        FROM ({union_sql})
    """)
    con.execute(f"""
        ALTER TABLE {stage_table} ADD COLUMN rec_id VARCHAR;
        UPDATE {stage_table} SET rec_id = 'node_{node_type}_' || md5(source_record_id);

        ALTER TABLE {stage_table} ADD COLUMN equipment_id VARCHAR;
        UPDATE {stage_table} SET equipment_id = e.equipment_id
        FROM equipment_master e
        WHERE {stage_table}.refinery_unit = e.refinery_unit
          AND norm_code({stage_table}.equipment_raw) = norm_code(e.equipment_code_clean)
          AND {stage_table}.equipment_raw IS NOT NULL;
    """)
    con.execute(f"""
        INSERT INTO node_raw
        SELECT rec_id, '{node_type}', source_record_id,
               coalesce(nullif(status_operation,''), '{domain_label} Readiness'), 'readiness_operation',
               json_object(
                   'refinery_unit', refinery_unit,
                   'equipment_raw', equipment_raw,
                   'period_date', period_date,
                   'status_operation', status_operation,
                   'status_item', status_item,
                   'remark', remark,
                   'rtl', rtl
               ),
               source_file, source_sheet, source_row, source_record_id
        FROM {stage_table} WHERE rec_id IS NOT NULL
    """)
    rel_type = f"EQUIPMENT_HAS_{node_type.upper()}"
    ru_rel_type = f"REFINERY_UNIT_HAS_{node_type.upper()}"
    con.execute(f"""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT equipment_id, rec_id, '{rel_type}',
               'readiness_operation', 1.0, 'ru_and_equipment_exact', false,
               json_object('match_token', equipment_raw),
               source_file, source_sheet, source_row, source_record_id
        FROM {stage_table} WHERE equipment_id IS NOT NULL AND rec_id IS NOT NULL
    """)
    con.execute(f"""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT r.refinery_unit_id, s.rec_id, '{ru_rel_type}',
               'readiness_operation', 1.0, 'refinery_unit_direct', false,
               json_object('refinery_unit', s.refinery_unit),
               s.source_file, s.source_sheet, s.source_row, s.source_record_id
        FROM {stage_table} s
        JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
        WHERE s.rec_id IS NOT NULL
    """)


def _build_rotor_nodes(con: duckdb.DuckDBPyConnection, views: list[str]) -> None:
    if not views:
        return
    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    ru_expr = f"ru_normalize({_cs(c, 'refinery_unit','ru','kilang', default='NULL')})"
    con.execute(f"""
        CREATE TABLE rotor_stage AS
        SELECT
            {_cs(c, 'equipment')} AS equipment_raw,
            {_cs(c, 'bulan','month','period', cast=True)} AS bulan,
            {_cs(c, 'program','program_name')} AS program,
            {_cs(c, 'brand')} AS brand,
            {_cs(c, 'status_readiness_spare_rotor','readiness_rotor','spare_rotor')} AS status_readiness,
            {_cs(c, 'status_workplan','status')} AS status_workplan,
            {_cs(c, 'detail_status_workplan','detail_status')} AS detail_status_workplan,
            {_cs(c, 'keterangan','remark','catatan')} AS keterangan,
            {_cs(c, 'action_plan_category','action_plan')} AS action_plan_category,
            {_cs(c, 'external_resource','resource')} AS external_resource,
            {_cs(c, 'no_irkap')} AS no_irkap,
            {_cs(c, 'finish_date_eksekusi','finish_date', cast=True)} AS finish_date,
            {_cs(c, 'last_update', cast=True)} AS last_update,
            coalesce({ru_expr}, ru_from_filename(_input_source_file)) AS refinery_unit,
            _input_source_file AS source_file,
            _input_source_sheet AS source_sheet,
            _source_row AS source_row,
            'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
        FROM ({union_sql})
        WHERE nullif(trim({_cs(c, 'equipment')}), '') IS NOT NULL
    """)
    con.execute("""
        ALTER TABLE rotor_stage ADD COLUMN rotor_id VARCHAR;
        UPDATE rotor_stage SET rotor_id = 'node_rotor_' || md5(source_record_id);

        ALTER TABLE rotor_stage ADD COLUMN equipment_id VARCHAR;
        UPDATE rotor_stage SET equipment_id = e.equipment_id
        FROM equipment_master e
        WHERE rotor_stage.refinery_unit = e.refinery_unit
          AND norm_code(rotor_stage.equipment_raw) = norm_code(e.equipment_code_clean);
    """)
    con.execute("""
        INSERT INTO node_raw
        SELECT rotor_id, 'rotor', source_record_id,
               coalesce(nullif(equipment_raw,''), 'Rotor'), 'rotor',
               json_object(
                   'refinery_unit', refinery_unit,
                   'equipment_raw', equipment_raw,
                   'bulan', bulan,
                   'program', program,
                   'brand', brand,
                   'status_readiness', status_readiness,
                   'status_workplan', status_workplan,
                   'detail_status_workplan', detail_status_workplan,
                   'keterangan', keterangan,
                   'action_plan_category', action_plan_category,
                   'external_resource', external_resource,
                   'no_irkap', no_irkap,
                   'finish_date', finish_date,
                   'last_update', last_update
               ),
               source_file, source_sheet, source_row, source_record_id
        FROM rotor_stage WHERE rotor_id IS NOT NULL
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT equipment_id, rotor_id, 'EQUIPMENT_HAS_ROTOR',
               'rotor', 1.0, 'ru_and_equipment_exact', false,
               json_object('match_token', equipment_raw),
               source_file, source_sheet, source_row, source_record_id
        FROM rotor_stage WHERE equipment_id IS NOT NULL AND rotor_id IS NOT NULL
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT r.refinery_unit_id, s.rotor_id, 'REFINERY_UNIT_HAS_ROTOR',
               'rotor', 1.0, 'refinery_unit_direct', false,
               json_object('refinery_unit', s.refinery_unit),
               s.source_file, s.source_sheet, s.source_row, s.source_record_id
        FROM rotor_stage s
        JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
        WHERE s.rotor_id IS NOT NULL
    """)
    con.execute("""
        INSERT INTO unmatched_raw
        SELECT equipment_raw, 'equipment', 'rotor', source_file, source_sheet, source_row,
               'Equipment tidak ditemukan di master'
        FROM rotor_stage
        WHERE equipment_id IS NULL AND nullif(trim(equipment_raw),'') IS NOT NULL
    """)


def _build_cross_domain_relationships(con: duckdb.DuckDBPyConnection) -> None:
    """Hubungkan antar domain via equipment yang sama — hanya rantai yang logis secara operasional."""
    existing = {r[0] for r in con.execute("SHOW TABLES").fetchall()}

    def _cross(src_table: str, src_id: str, tgt_table: str, tgt_id: str,
                rel_type: str, domain: str) -> None:
        if src_table not in existing or tgt_table not in existing:
            return
        con.execute(f"""
            INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
                domain, confidence, match_rule, is_candidate, properties_json,
                source_file, source_sheet, source_row, source_record_id)
            SELECT DISTINCT s.{src_id}, t.{tgt_id}, '{rel_type}',
                   '{domain}', 1.0, 'equipment_same_match', false,
                   json_object('equipment_id', s.equipment_id),
                   s.source_file, s.source_sheet, s.source_row, s.source_record_id
            FROM {src_table} s
            JOIN {tgt_table} t ON s.equipment_id = t.equipment_id
            WHERE s.{src_id} IS NOT NULL AND t.{tgt_id} IS NOT NULL
              AND s.equipment_id IS NOT NULL
        """)

    # Rantai 1: Critical Equipment → Bad Actor → (RCPS via no_irkap sudah ada)
    _cross('crit_eq_stage', 'ce_id', 'bad_actor_stage', 'bad_actor_id',
           'CRITICAL_EQUIPMENT_HAS_BAD_ACTOR', 'critical_equipment')

    # Rantai 2: Zero Clamp → Inspection & Pipeline Inspection
    _cross('zero_clamp_stage', 'zc_id', 'inspection_stage', 'inspection_id',
           'ZERO_CLAMP_HAS_INSPECTION', 'zero_clamp')
    _cross('zero_clamp_stage', 'zc_id', 'pipeline_insp_stage', 'pi_id',
           'ZERO_CLAMP_HAS_PIPELINE_INSPECTION', 'zero_clamp')

    # Rantai 3: Power & Steam → Monitoring Operasi
    if 'power_steam_stage' in existing and 'mon_operasi_stage' in existing:
        con.execute("""
            INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
                domain, confidence, match_rule, is_candidate, properties_json,
                source_file, source_sheet, source_row, source_record_id)
            SELECT DISTINCT ps.ps_id, mo.mo_id, 'POWER_STEAM_HAS_MONITORING_OPERASI',
                   'power_steam', 1.0, 'equipment_same_match', false,
                   json_object('equipment_id', ps.equipment_id),
                   ps.source_file, ps.source_sheet, ps.source_row, ps.source_record_id
            FROM power_steam_stage ps
            JOIN mon_operasi_stage mo
              ON ps.equipment_id = mo.eq_process_id
              OR ps.equipment_id = mo.eq_sts_id
            WHERE ps.ps_id IS NOT NULL AND mo.mo_id IS NOT NULL
              AND ps.equipment_id IS NOT NULL
        """)

    # Rantai 5: Reliability → ICU Issue (via equipment)
    if 'reliability_stage' in existing and 'icu_stage' in existing:
        con.execute("""
            INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
                domain, confidence, match_rule, is_candidate, properties_json,
                source_file, source_sheet, source_row, source_record_id)
            SELECT DISTINCT r.obs_id, i.issue_id, 'RELIABILITY_HAS_ICU_ISSUE',
                   'reliability', 1.0, 'equipment_same_match', false,
                   json_object('equipment_id', r.equipment_id),
                   r.source_file, r.source_sheet, r.source_row, r.source_record_id
            FROM reliability_stage r
            JOIN icu_stage i ON r.equipment_id = i.equipment_id
            WHERE r.obs_id IS NOT NULL AND i.issue_id IS NOT NULL
              AND r.equipment_id IS NOT NULL
        """)


def _build_bad_actor_nodes(con: duckdb.DuckDBPyConnection, views: list[str]) -> None:
    if not views:
        return
    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    ru_expr = f"ru_normalize({_cs(c, 'refinery_unit','ru','kilang', default='NULL')})"
    tag_expr = _cs(c, 'tag_number', 'tag_no', 'equipment')
    con.execute(f"""
        CREATE TABLE bad_actor_stage AS
        SELECT
            {tag_expr} AS tag_raw,
            {_cs(c, 'problem','masalah','highlight_issue')} AS problem,
            {_cs(c, 'status')} AS status,
            {_cs(c, 'action_plan','corrective_action')} AS action_plan,
            {_cs(c, 'category_action_plan','action_plan_category')} AS category,
            {_cs(c, 'progress')} AS progress,
            {_cs(c, 'target_date','target')} AS target_date,
            {_cs(c, 'no_irkap')} AS no_irkap,
            {_cs(c, 'action_plan_remark','remark')} AS remark,
            {_cs(c, 'unit_proses','unit')} AS unit_proses,
            coalesce({ru_expr}, ru_from_filename(_input_source_file)) AS refinery_unit,
            _input_source_file AS source_file,
            _input_source_sheet AS source_sheet,
            _source_row AS source_row,
            'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
        FROM ({union_sql})
        WHERE nullif(trim({tag_expr}), '') IS NOT NULL
    """)
    con.execute("""
        ALTER TABLE bad_actor_stage ADD COLUMN bad_actor_id VARCHAR;
        UPDATE bad_actor_stage
        SET bad_actor_id = 'node_bad_actor_' || md5(coalesce(refinery_unit,'UNKNOWN') || '|' || norm_code(tag_raw) || '|' || source_record_id);
        ALTER TABLE bad_actor_stage ADD COLUMN equipment_id VARCHAR;
        UPDATE bad_actor_stage SET equipment_id = e.equipment_id
        FROM equipment_master e
        WHERE bad_actor_stage.refinery_unit = e.refinery_unit
          AND norm_code(bad_actor_stage.tag_raw) = norm_code(e.equipment_code_clean);
    """)
    con.execute("""
        INSERT INTO node_raw
        SELECT bad_actor_id, 'bad_actor', tag_raw,
               coalesce(nullif(left(problem,90),''), tag_raw), 'bad_actor',
               json_object('refinery_unit', refinery_unit, 'tag_raw', tag_raw,
                   'problem', problem, 'status', status, 'action_plan', action_plan,
                   'category', category, 'progress', progress, 'target_date', target_date,
                   'no_irkap', no_irkap, 'remark', remark, 'unit_proses', unit_proses),
               source_file, source_sheet, source_row, source_record_id
        FROM bad_actor_stage WHERE bad_actor_id IS NOT NULL
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT equipment_id, bad_actor_id, 'EQUIPMENT_HAS_BAD_ACTOR',
               'bad_actor', 1.0, 'equipment_exact', false,
               json_object('tag_raw', tag_raw),
               source_file, source_sheet, source_row, source_record_id
        FROM bad_actor_stage WHERE equipment_id IS NOT NULL AND bad_actor_id IS NOT NULL
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT DISTINCT r.refinery_unit_id, s.bad_actor_id, 'REFINERY_UNIT_HAS_BAD_ACTOR',
               'bad_actor', 1.0, 'refinery_unit_direct', false,
               json_object('refinery_unit', s.refinery_unit),
               s.source_file, s.source_sheet, s.source_row, s.source_record_id
        FROM bad_actor_stage s
        JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
        WHERE s.bad_actor_id IS NOT NULL
    """)


def _build_zero_clamp_nodes(con: duckdb.DuckDBPyConnection, views: list[str]) -> None:
    if not views:
        return
    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    ru_expr = f"ru_normalize({_cs(c, 'refinery_unit','ru','kilang', default='NULL')})"
    tag_expr = _cs(c, 'tag_no_ln', 'tag_number', 'tag_no')
    eq_expr = _cs(c, 'equipment', 'tag_no_ln', 'tag_number')
    con.execute(f"""
        CREATE TABLE zero_clamp_stage AS
        SELECT
            {tag_expr} AS tag_raw,
            {eq_expr} AS equipment_raw,
            {_cs(c, 'description','deskripsi')} AS description,
            {_cs(c, 'services','service')} AS services,
            {_cs(c, 'type_damage','jenis_kerusakan')} AS type_damage,
            {_cs(c, 'posisi')} AS posisi,
            {_cs(c, 'type_perbaikan','jenis_perbaikan')} AS type_perbaikan,
            {_cs(c, 'tanggal_dipasang','tgl_pasang')} AS tanggal_dipasang,
            {_cs(c, 'tanggal_dilepas','tgl_lepas')} AS tanggal_dilepas,
            {_cs(c, 'tanggal_rencana_perbaikan')} AS tanggal_rencana,
            {_cs(c, 'no_irkap')} AS no_irkap,
            {_cs(c, 'status')} AS status,
            {_cs(c, 'remarks','remark')} AS remarks,
            {_cs(c, 'area')} AS area,
            {_cs(c, 'unit')} AS unit,
            coalesce({ru_expr}, ru_from_filename(_input_source_file)) AS refinery_unit,
            _input_source_file AS source_file,
            _input_source_sheet AS source_sheet,
            _source_row AS source_row,
            'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
        FROM ({union_sql})
        WHERE nullif(trim({tag_expr}), '') IS NOT NULL
    """)
    con.execute("""
        ALTER TABLE zero_clamp_stage ADD COLUMN zc_id VARCHAR;
        UPDATE zero_clamp_stage
        SET zc_id = 'node_zero_clamp_' || md5(coalesce(refinery_unit,'UNKNOWN') || '|' || tag_raw || '|' || source_record_id);
        ALTER TABLE zero_clamp_stage ADD COLUMN equipment_id VARCHAR;
        UPDATE zero_clamp_stage SET equipment_id = e.equipment_id
        FROM equipment_master e
        WHERE zero_clamp_stage.refinery_unit = e.refinery_unit
          AND norm_code(zero_clamp_stage.equipment_raw) = norm_code(e.equipment_code_clean);
    """)
    con.execute("""
        INSERT INTO node_raw
        SELECT zc_id, 'zero_clamp', tag_raw,
               coalesce(nullif(description,''), tag_raw), 'zero_clamp',
               json_object('refinery_unit', refinery_unit, 'tag_raw', tag_raw,
                   'services', services, 'type_damage', type_damage, 'posisi', posisi,
                   'type_perbaikan', type_perbaikan, 'tanggal_dipasang', tanggal_dipasang,
                   'tanggal_dilepas', tanggal_dilepas, 'tanggal_rencana', tanggal_rencana,
                   'no_irkap', no_irkap, 'status', status, 'remarks', remarks,
                   'area', area, 'unit', unit),
               source_file, source_sheet, source_row, source_record_id
        FROM zero_clamp_stage WHERE zc_id IS NOT NULL
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT equipment_id, zc_id, 'EQUIPMENT_HAS_ZERO_CLAMP',
               'zero_clamp', 1.0, 'equipment_exact', false,
               json_object('tag_raw', tag_raw),
               source_file, source_sheet, source_row, source_record_id
        FROM zero_clamp_stage WHERE equipment_id IS NOT NULL AND zc_id IS NOT NULL
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT DISTINCT r.refinery_unit_id, s.zc_id, 'REFINERY_UNIT_HAS_ZERO_CLAMP',
               'zero_clamp', 1.0, 'refinery_unit_direct', false,
               json_object('refinery_unit', s.refinery_unit),
               s.source_file, s.source_sheet, s.source_row, s.source_record_id
        FROM zero_clamp_stage s
        JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
        WHERE s.zc_id IS NOT NULL
    """)


def _build_paf_nodes(con: duckdb.DuckDBPyConnection,
                     paf_views: list[str], issue_views: list[str]) -> None:
    if paf_views:
        c = _union_cols(con, paf_views)
        union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in paf_views)
        ru_expr = f"ru_normalize({_cs(c, 'ru','refinery_unit','kilang','ru2', default='NULL')})"
        con.execute(f"""
            CREATE TABLE paf_stage AS
            SELECT
                {_cs(c, 'type','tipe')} AS type_unit,
                {_cs(c, 'target_realisasi')} AS target_realisasi,
                {_cs(c, 'color','warna')} AS color,
                {_cs(c, 'value','nilai', cast=True, default="'0'")} AS value,
                {_cs(c, 'plan_unplan')} AS plan_unplan,
                {_cs(c, 'type2')} AS type2,
                {_cs(c, 'month','bulan')} AS month,
                {_cs(c, 'month_update')} AS month_update,
                {_cs(c, 'target')} AS target,
                {_cs(c, 'code_current', cast=True, default="'0'")} AS code_current,
                coalesce({ru_expr}, ru_from_filename(_input_source_file)) AS refinery_unit,
                _input_source_file AS source_file,
                _input_source_sheet AS source_sheet,
                _source_row AS source_row,
                'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
            FROM ({union_sql})
            WHERE nullif(trim({_cs(c, 'type','tipe')}), '') IS NOT NULL
        """)
        con.execute("""
            ALTER TABLE paf_stage ADD COLUMN paf_id VARCHAR;
            UPDATE paf_stage
            SET paf_id = 'node_paf_' || md5(coalesce(refinery_unit,'UNKNOWN') || '|' || coalesce(type_unit,'') || '|' || coalesce(month_update,'') || '|' || source_record_id);
        """)
        con.execute("""
            INSERT INTO node_raw
            SELECT paf_id, 'paf', coalesce(refinery_unit,'UNKNOWN') || '|' || coalesce(type_unit,'') || '|' || coalesce(month_update,''),
                   coalesce(type_unit, 'PAF'), 'paf',
                   json_object('refinery_unit', refinery_unit, 'type_unit', type_unit,
                       'target_realisasi', target_realisasi, 'color', color, 'value', value,
                       'plan_unplan', plan_unplan, 'type2', type2, 'month', month,
                       'month_update', month_update, 'target', target, 'code_current', code_current),
                   source_file, source_sheet, source_row, source_record_id
            FROM paf_stage WHERE paf_id IS NOT NULL
        """)
        con.execute("""
            INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
                domain, confidence, match_rule, is_candidate, properties_json,
                source_file, source_sheet, source_row, source_record_id)
            SELECT DISTINCT r.refinery_unit_id, s.paf_id, 'REFINERY_UNIT_HAS_PAF',
                   'paf', 1.0, 'refinery_unit_direct', false,
                   json_object('refinery_unit', s.refinery_unit),
                   s.source_file, s.source_sheet, s.source_row, s.source_record_id
            FROM paf_stage s
            JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
            WHERE s.paf_id IS NOT NULL
        """)

    if issue_views:
        c2 = _union_cols(con, issue_views)
        union_sql2 = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in issue_views)
        ru_expr2 = f"ru_normalize({_cs(c2, 'ru','refinery_unit','kilang', default='NULL')})"
        con.execute(f"""
            CREATE TABLE paf_issue_stage AS
            SELECT
                {_cs(c2, 'type','tipe')} AS type_unit,
                {_cs(c2, 'issue','permasalahan','masalah')} AS issue,
                {_cs(c2, 'date','tanggal', cast=True)} AS issue_date,
                {_cs(c2, 'month_update')} AS month_update,
                {_cs(c2, 'code_current', cast=True, default="'0'")} AS code_current,
                coalesce({ru_expr2}, ru_from_filename(_input_source_file)) AS refinery_unit,
                _input_source_file AS source_file,
                _input_source_sheet AS source_sheet,
                _source_row AS source_row,
                'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
            FROM ({union_sql2})
            WHERE nullif(trim({_cs(c2, 'issue','permasalahan','masalah')}), '') IS NOT NULL
        """)
        con.execute("""
            ALTER TABLE paf_issue_stage ADD COLUMN paf_issue_id VARCHAR;
            UPDATE paf_issue_stage
            SET paf_issue_id = 'node_paf_issue_' || md5(source_record_id);
        """)
        con.execute("""
            INSERT INTO node_raw
            SELECT paf_issue_id, 'paf_issue', source_record_id,
                   coalesce(nullif(left(issue,90),''), 'PAF Issue'), 'paf',
                   json_object('refinery_unit', refinery_unit, 'type_unit', type_unit,
                       'issue', issue, 'issue_date', issue_date, 'month_update', month_update,
                       'code_current', code_current),
                   source_file, source_sheet, source_row, source_record_id
            FROM paf_issue_stage WHERE paf_issue_id IS NOT NULL
        """)
        con.execute("""
            INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
                domain, confidence, match_rule, is_candidate, properties_json,
                source_file, source_sheet, source_row, source_record_id)
            SELECT DISTINCT r.refinery_unit_id, s.paf_issue_id, 'REFINERY_UNIT_HAS_PAF_ISSUE',
                   'paf', 1.0, 'refinery_unit_direct', false,
                   json_object('refinery_unit', s.refinery_unit),
                   s.source_file, s.source_sheet, s.source_row, s.source_record_id
            FROM paf_issue_stage s
            JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
            WHERE s.paf_issue_id IS NOT NULL
        """)


def _build_atg_nodes(con: duckdb.DuckDBPyConnection,
                     atg_views: list[str], program_views: list[str]) -> None:
    if atg_views:
        c = _union_cols(con, atg_views)
        union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in atg_views)
        ru_expr = f"ru_normalize({_cs(c, 'refinery_unit','ru','kilang', default='NULL')})"
        eq_tangki = _cs(c, 'equipment_tangki', 'tag_no_tangki')
        eq_atg = _cs(c, 'equipment_atg', 'tag_no_atg')
        con.execute(f"""
            CREATE TABLE atg_stage AS
            SELECT
                {_cs(c, 'tag_no_tangki','tag_tangki')} AS tag_tangki,
                {_cs(c, 'tag_no_atg','tag_atg')} AS tag_atg,
                {_cs(c, 'status_atg','status')} AS status_atg,
                {_cs(c, 'status_interkoneksi_atg','status_interkoneksi')} AS status_interkoneksi,
                {_cs(c, 'cert_no_atg','cert_no')} AS cert_no,
                {_cs(c, 'date_expired_atg','date_expired')} AS date_expired,
                {_cs(c, 'remark','keterangan')} AS remark,
                {_cs(c, 'rtl')} AS rtl,
                {_cs(c, 'status_rtl')} AS status_rtl,
                {_cs(c, 'no_irkap')} AS no_irkap,
                {_cs(c, 'month_update')} AS month_update,
                {eq_tangki} AS equipment_tangki_raw,
                {eq_atg} AS equipment_atg_raw,
                coalesce({ru_expr}, ru_from_filename(_input_source_file)) AS refinery_unit,
                _input_source_file AS source_file,
                _input_source_sheet AS source_sheet,
                _source_row AS source_row,
                'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
            FROM ({union_sql})
            WHERE nullif(trim({_cs(c, 'tag_no_tangki','tag_tangki')}), '') IS NOT NULL
        """)
        con.execute("""
            ALTER TABLE atg_stage ADD COLUMN atg_id VARCHAR;
            UPDATE atg_stage
            SET atg_id = 'node_atg_' || md5(coalesce(refinery_unit,'UNKNOWN') || '|' || coalesce(tag_tangki,'') || '|' || coalesce(tag_atg,''));
            ALTER TABLE atg_stage ADD COLUMN eq_tangki_id VARCHAR;
            UPDATE atg_stage SET eq_tangki_id = e.equipment_id
            FROM equipment_master e
            WHERE atg_stage.refinery_unit = e.refinery_unit
              AND norm_code(atg_stage.equipment_tangki_raw) = norm_code(e.equipment_code_clean);
            ALTER TABLE atg_stage ADD COLUMN eq_atg_id VARCHAR;
            UPDATE atg_stage SET eq_atg_id = e.equipment_id
            FROM equipment_master e
            WHERE atg_stage.refinery_unit = e.refinery_unit
              AND norm_code(atg_stage.equipment_atg_raw) = norm_code(e.equipment_code_clean);
        """)
        con.execute("""
            INSERT INTO node_raw
            SELECT DISTINCT ON (atg_id)
                   atg_id, 'atg', coalesce(tag_tangki, source_record_id),
                   coalesce(tag_tangki, 'ATG'), 'atg',
                   json_object('refinery_unit', refinery_unit, 'tag_tangki', tag_tangki,
                       'tag_atg', tag_atg, 'status_atg', status_atg,
                       'status_interkoneksi', status_interkoneksi, 'cert_no', cert_no,
                       'date_expired', date_expired, 'remark', remark, 'rtl', rtl,
                       'status_rtl', status_rtl, 'no_irkap', no_irkap),
                   source_file, source_sheet, source_row, source_record_id
            FROM atg_stage WHERE atg_id IS NOT NULL ORDER BY atg_id, source_row
        """)
        for eq_col, rel_type in [('eq_tangki_id', 'EQUIPMENT_HAS_ATG'), ('eq_atg_id', 'EQUIPMENT_HAS_ATG')]:
            con.execute(f"""
                INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
                    domain, confidence, match_rule, is_candidate, properties_json,
                    source_file, source_sheet, source_row, source_record_id)
                SELECT {eq_col}, atg_id, '{rel_type}',
                       'atg', 1.0, 'equipment_exact', false,
                       json_object('tag_tangki', tag_tangki),
                       source_file, source_sheet, source_row, source_record_id
                FROM atg_stage WHERE {eq_col} IS NOT NULL AND atg_id IS NOT NULL
            """)
        con.execute("""
            INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
                domain, confidence, match_rule, is_candidate, properties_json,
                source_file, source_sheet, source_row, source_record_id)
            SELECT DISTINCT r.refinery_unit_id, s.atg_id, 'REFINERY_UNIT_HAS_ATG',
                   'atg', 1.0, 'refinery_unit_direct', false,
                   json_object('refinery_unit', s.refinery_unit),
                   s.source_file, s.source_sheet, s.source_row, s.source_record_id
            FROM atg_stage s
            JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
            WHERE s.atg_id IS NOT NULL
        """)

    if program_views:
        c2 = _union_cols(con, program_views)
        union_sql2 = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in program_views)
        ru_expr2 = f"ru_normalize({_cs(c2, 'refinery_unit','ru','kilang', default='NULL')})"
        con.execute(f"""
            CREATE TABLE atg_program_stage AS
            SELECT
                {_cs(c2, 'type','tipe')} AS type_program,
                {_cs(c2, 'atg_eksisting')} AS atg_eksisting,
                {_cs(c2, 'program_2024','program_kerja','program')} AS program_name,
                {_cs(c2, 'prokja','progress')} AS prokja,
                {_cs(c2, 'action_plan_category','kategori')} AS category,
                {_cs(c2, 'target')} AS target,
                {_cs(c2, 'month_update')} AS month_update,
                coalesce({ru_expr2}, ru_from_filename(_input_source_file)) AS refinery_unit,
                _input_source_file AS source_file,
                _input_source_sheet AS source_sheet,
                _source_row AS source_row,
                'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
            FROM ({union_sql2})
            WHERE nullif(trim({_cs(c2, 'program_2024','program_kerja','program')}), '') IS NOT NULL
        """)
        con.execute("""
            ALTER TABLE atg_program_stage ADD COLUMN prog_id VARCHAR;
            UPDATE atg_program_stage
            SET prog_id = 'node_atg_program_' || md5(source_record_id);
        """)
        con.execute("""
            INSERT INTO node_raw
            SELECT prog_id, 'atg_program', source_record_id,
                   coalesce(nullif(left(program_name,90),''), 'ATG Program'), 'atg',
                   json_object('refinery_unit', refinery_unit, 'type_program', type_program,
                       'atg_eksisting', atg_eksisting, 'program_name', program_name,
                       'prokja', prokja, 'category', category, 'target', target,
                       'month_update', month_update),
                   source_file, source_sheet, source_row, source_record_id
            FROM atg_program_stage WHERE prog_id IS NOT NULL
        """)
        con.execute("""
            INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
                domain, confidence, match_rule, is_candidate, properties_json,
                source_file, source_sheet, source_row, source_record_id)
            SELECT DISTINCT r.refinery_unit_id, s.prog_id, 'REFINERY_UNIT_HAS_ATG_PROGRAM',
                   'atg', 1.0, 'refinery_unit_direct', false,
                   json_object('refinery_unit', s.refinery_unit),
                   s.source_file, s.source_sheet, s.source_row, s.source_record_id
            FROM atg_program_stage s
            JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
            WHERE s.prog_id IS NOT NULL
        """)


def _build_pipeline_inspection_nodes(con: duckdb.DuckDBPyConnection, views: list[str]) -> None:
    if not views:
        return
    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    ru_expr = f"ru_normalize({_cs(c, 'refinery_unit','ru','kilang', default='NULL')})"
    tag_expr = _cs(c, 'tag_number', 'tag_no', 'tag_no_ln')
    eq_expr = _cs(c, 'equipment', 'tag_number', 'tag_no')
    con.execute(f"""
        CREATE TABLE pipeline_insp_stage AS
        SELECT
            {tag_expr} AS tag_raw,
            {eq_expr} AS equipment_raw,
            {_cs(c, 'last_inspection_date','last_inspection', cast=True)} AS last_inspection,
            {_cs(c, 'next_inspection_date','next_inspection', cast=True)} AS next_inspection,
            {_cs(c, 'fluida_service','service','services')} AS fluida_service,
            {_cs(c, 'nps')} AS nps,
            {_cs(c, 'from_location','from_loc')} AS from_location,
            {_cs(c, 'to_location','to_loc')} AS to_location,
            {_cs(c, 'last_measured_thickness', cast=True, default="'0'")} AS thickness,
            {_cs(c, 'rem_life_years', cast=True, default="'0'")} AS rem_life,
            {_cs(c, 'jumlah_temporary_repair', cast=True, default="'0'")} AS temp_repair,
            {_cs(c, 'remarks','remark')} AS remarks,
            {_cs(c, 'area')} AS area,
            {_cs(c, 'unit')} AS unit,
            coalesce({ru_expr}, ru_from_filename(_input_source_file)) AS refinery_unit,
            _input_source_file AS source_file,
            _input_source_sheet AS source_sheet,
            _source_row AS source_row,
            'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
        FROM ({union_sql})
        WHERE nullif(trim({tag_expr}), '') IS NOT NULL
    """)
    con.execute("""
        ALTER TABLE pipeline_insp_stage ADD COLUMN pi_id VARCHAR;
        UPDATE pipeline_insp_stage
        SET pi_id = 'node_pipeline_inspection_' || md5(coalesce(refinery_unit,'UNKNOWN') || '|' || tag_raw);
        ALTER TABLE pipeline_insp_stage ADD COLUMN equipment_id VARCHAR;
        UPDATE pipeline_insp_stage SET equipment_id = e.equipment_id
        FROM equipment_master e
        WHERE pipeline_insp_stage.refinery_unit = e.refinery_unit
          AND norm_code(pipeline_insp_stage.equipment_raw) = norm_code(e.equipment_code_clean);
    """)
    con.execute("""
        INSERT INTO node_raw
        SELECT DISTINCT ON (pi_id)
               pi_id, 'pipeline_inspection', tag_raw,
               coalesce(nullif(tag_raw,''), 'Pipeline'), 'pipeline_inspection',
               json_object('refinery_unit', refinery_unit, 'tag_raw', tag_raw,
                   'last_inspection', last_inspection, 'next_inspection', next_inspection,
                   'fluida_service', fluida_service, 'nps', nps,
                   'from_location', from_location, 'to_location', to_location,
                   'thickness', thickness, 'rem_life', rem_life,
                   'temp_repair', temp_repair, 'remarks', remarks,
                   'area', area, 'unit', unit),
               source_file, source_sheet, source_row, source_record_id
        FROM pipeline_insp_stage WHERE pi_id IS NOT NULL ORDER BY pi_id, source_row
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT equipment_id, pi_id, 'EQUIPMENT_HAS_PIPELINE_INSPECTION',
               'pipeline_inspection', 1.0, 'equipment_exact', false,
               json_object('tag_raw', tag_raw),
               source_file, source_sheet, source_row, source_record_id
        FROM pipeline_insp_stage WHERE equipment_id IS NOT NULL AND pi_id IS NOT NULL
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT DISTINCT r.refinery_unit_id, s.pi_id, 'REFINERY_UNIT_HAS_PIPELINE_INSPECTION',
               'pipeline_inspection', 1.0, 'refinery_unit_direct', false,
               json_object('refinery_unit', s.refinery_unit),
               s.source_file, s.source_sheet, s.source_row, s.source_record_id
        FROM pipeline_insp_stage s
        JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
        WHERE s.pi_id IS NOT NULL
    """)


def _build_tkdn_nodes(con: duckdb.DuckDBPyConnection, views: list[str]) -> None:
    if not views:
        return
    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    ru_expr = f"ru_normalize({_cs(c, 'refinery_unit','ru','kilang', default='NULL')})"
    con.execute(f"""
        CREATE TABLE tkdn_stage AS
        SELECT
            {_cs(c, 'bulan','month','periode_bulan')} AS bulan,
            {_cs(c, 'tahun','year', cast=True)} AS tahun,
            {_cs(c, 'nominal', cast=True, default="'0'")} AS nominal,
            {_cs(c, 'kdn', cast=True, default="'0'")} AS kdn,
            {_cs(c, 'persentase','percentage', cast=True, default="'0'")} AS persentase,
            coalesce({ru_expr}, ru_from_filename(_input_source_file)) AS refinery_unit,
            _input_source_file AS source_file,
            _input_source_sheet AS source_sheet,
            _source_row AS source_row,
            'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
        FROM ({union_sql})
        WHERE nullif(trim({_cs(c, 'bulan','month','periode_bulan')}), '') IS NOT NULL
    """)
    con.execute("""
        ALTER TABLE tkdn_stage ADD COLUMN tkdn_id VARCHAR;
        UPDATE tkdn_stage
        SET tkdn_id = 'node_tkdn_' || md5(coalesce(refinery_unit,'UNKNOWN') || '|' || coalesce(bulan,'') || '|' || coalesce(tahun,''));
    """)
    con.execute("""
        INSERT INTO node_raw
        SELECT DISTINCT ON (tkdn_id)
               tkdn_id, 'tkdn', coalesce(refinery_unit,'UNKNOWN') || '|' || coalesce(bulan,'') || '|' || coalesce(tahun,''),
               coalesce(bulan,'') || ' ' || coalesce(tahun,''), 'tkdn',
               json_object('refinery_unit', refinery_unit, 'bulan', bulan, 'tahun', tahun,
                   'nominal', nominal, 'kdn', kdn, 'persentase', persentase),
               source_file, source_sheet, source_row, source_record_id
        FROM tkdn_stage WHERE tkdn_id IS NOT NULL ORDER BY tkdn_id, source_row
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT DISTINCT r.refinery_unit_id, s.tkdn_id, 'REFINERY_UNIT_HAS_TKDN',
               'tkdn', 1.0, 'refinery_unit_direct', false,
               json_object('refinery_unit', s.refinery_unit),
               s.source_file, s.source_sheet, s.source_row, s.source_record_id
        FROM tkdn_stage s
        JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
        WHERE s.tkdn_id IS NOT NULL
    """)


def _build_monitoring_operasi_nodes(con: duckdb.DuckDBPyConnection, views: list[str]) -> None:
    if not views:
        return
    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    ru_expr = f"ru_normalize({_cs(c, 'refinery_unit','ru','kilang', default='NULL')})"
    eq_process = _cs(c, 'equipment_process')
    eq_sts = _cs(c, 'equipment_sts')
    con.execute(f"""
        CREATE TABLE mon_operasi_stage AS
        SELECT
            {_cs(c, 'unit_proses','unit_proses')} AS unit_proses,
            {_cs(c, 'unit')} AS unit,
            {_cs(c, 'unit_measurement')} AS unit_measurement,
            {_cs(c, 'design', cast=True, default="'0'")} AS design,
            {_cs(c, 'minimal_capacity', cast=True, default="'0'")} AS minimal_capacity,
            {_cs(c, 'plant_readiness')} AS plant_readiness,
            {_cs(c, 'type_limitasi_process')} AS type_limitasi_process,
            {eq_process} AS equipment_process_raw,
            {_cs(c, 'limitasi_alert_process')} AS limitasi_alert_process,
            {_cs(c, 'mitigasi_process','mitigasi_action')} AS mitigasi_process,
            {_cs(c, 'target_sts', cast=True, default="'0'")} AS target_sts,
            {_cs(c, 'actual', cast=True, default="'0'")} AS actual,
            {_cs(c, 'type_limitasi_sts')} AS type_limitasi_sts,
            {eq_sts} AS equipment_sts_raw,
            {_cs(c, 'limitasi_alert_sts')} AS limitasi_alert_sts,
            {_cs(c, 'month_update')} AS month_update,
            coalesce({ru_expr}, ru_from_filename(_input_source_file)) AS refinery_unit,
            _input_source_file AS source_file,
            _input_source_sheet AS source_sheet,
            _source_row AS source_row,
            'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
        FROM ({union_sql})
        WHERE nullif(trim({_cs(c, 'unit')}), '') IS NOT NULL
    """)
    con.execute("""
        ALTER TABLE mon_operasi_stage ADD COLUMN mo_id VARCHAR;
        UPDATE mon_operasi_stage
        SET mo_id = 'node_monitoring_operasi_' || md5(source_record_id);
        ALTER TABLE mon_operasi_stage ADD COLUMN eq_process_id VARCHAR;
        UPDATE mon_operasi_stage SET eq_process_id = e.equipment_id
        FROM equipment_master e
        WHERE mon_operasi_stage.refinery_unit = e.refinery_unit
          AND norm_code(mon_operasi_stage.equipment_process_raw) = norm_code(e.equipment_code_clean)
          AND nullif(trim(mon_operasi_stage.equipment_process_raw), '') IS NOT NULL;
        ALTER TABLE mon_operasi_stage ADD COLUMN eq_sts_id VARCHAR;
        UPDATE mon_operasi_stage SET eq_sts_id = e.equipment_id
        FROM equipment_master e
        WHERE mon_operasi_stage.refinery_unit = e.refinery_unit
          AND norm_code(mon_operasi_stage.equipment_sts_raw) = norm_code(e.equipment_code_clean)
          AND nullif(trim(mon_operasi_stage.equipment_sts_raw), '') IS NOT NULL;
    """)
    con.execute("""
        INSERT INTO node_raw
        SELECT mo_id, 'monitoring_operasi', source_record_id,
               coalesce(unit, 'Monitoring'), 'monitoring_operasi',
               json_object('refinery_unit', refinery_unit, 'unit_proses', unit_proses,
                   'unit', unit, 'unit_measurement', unit_measurement,
                   'design', design, 'minimal_capacity', minimal_capacity,
                   'plant_readiness', plant_readiness,
                   'type_limitasi_process', type_limitasi_process,
                   'limitasi_alert_process', limitasi_alert_process,
                   'target_sts', target_sts, 'actual', actual,
                   'type_limitasi_sts', type_limitasi_sts,
                   'limitasi_alert_sts', limitasi_alert_sts,
                   'month_update', month_update),
               source_file, source_sheet, source_row, source_record_id
        FROM mon_operasi_stage WHERE mo_id IS NOT NULL
    """)
    for eq_col, rel_type in [('eq_process_id', 'EQUIPMENT_HAS_MONITORING_OPERASI'),
                              ('eq_sts_id', 'EQUIPMENT_HAS_MONITORING_OPERASI')]:
        con.execute(f"""
            INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
                domain, confidence, match_rule, is_candidate, properties_json,
                source_file, source_sheet, source_row, source_record_id)
            SELECT {eq_col}, mo_id, '{rel_type}',
                   'monitoring_operasi', 1.0, 'equipment_exact', false,
                   json_object('unit', unit),
                   source_file, source_sheet, source_row, source_record_id
            FROM mon_operasi_stage WHERE {eq_col} IS NOT NULL AND mo_id IS NOT NULL
        """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT DISTINCT r.refinery_unit_id, s.mo_id, 'REFINERY_UNIT_HAS_MONITORING_OPERASI',
               'monitoring_operasi', 1.0, 'refinery_unit_direct', false,
               json_object('refinery_unit', s.refinery_unit),
               s.source_file, s.source_sheet, s.source_row, s.source_record_id
        FROM mon_operasi_stage s
        JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
        WHERE s.mo_id IS NOT NULL
    """)


def _build_power_steam_nodes(con: duckdb.DuckDBPyConnection, views: list[str]) -> None:
    if not views:
        return
    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    ru_expr = f"ru_normalize({_cs(c, 'refinery_unit','ru','kilang', default='NULL')})"
    eq_expr = _cs(c, 'equipment', 'tag_number', 'tag_no')
    con.execute(f"""
        CREATE TABLE power_steam_stage AS
        SELECT
            {_cs(c, 'type_equipment','tipe_equipment','type')} AS type_equipment,
            {eq_expr} AS equipment_raw,
            {_cs(c, 'status_operation','status_operasi')} AS status_operation,
            {_cs(c, 'status_n0','status_n')} AS status_n0,
            {_cs(c, 'unit_measurement')} AS unit_measurement,
            {_cs(c, 'desain','design', cast=True, default="'0'")} AS desain,
            {_cs(c, 'kapasitas_max','kapasitas', cast=True, default="'0'")} AS kapasitas_max,
            {_cs(c, 'average_actual','actual', cast=True, default="'0'")} AS average_actual,
            {_cs(c, 'remark','keterangan')} AS remark,
            {_cs(c, 'month_update','date_update')} AS month_update,
            coalesce({ru_expr}, ru_from_filename(_input_source_file)) AS refinery_unit,
            _input_source_file AS source_file,
            _input_source_sheet AS source_sheet,
            _source_row AS source_row,
            'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
        FROM ({union_sql})
        WHERE nullif(trim({eq_expr}), '') IS NOT NULL
    """)
    con.execute("""
        ALTER TABLE power_steam_stage ADD COLUMN ps_id VARCHAR;
        UPDATE power_steam_stage
        SET ps_id = 'node_power_steam_' || md5(coalesce(refinery_unit,'UNKNOWN') || '|' || equipment_raw || '|' || coalesce(month_update,''));
        ALTER TABLE power_steam_stage ADD COLUMN equipment_id VARCHAR;
        UPDATE power_steam_stage SET equipment_id = e.equipment_id
        FROM equipment_master e
        WHERE power_steam_stage.refinery_unit = e.refinery_unit
          AND norm_code(power_steam_stage.equipment_raw) = norm_code(e.equipment_code_clean);
    """)
    con.execute("""
        INSERT INTO node_raw
        SELECT DISTINCT ON (ps_id)
               ps_id, 'power_steam', equipment_raw,
               coalesce(nullif(equipment_raw,''), 'Power/Steam'), 'power_steam',
               json_object('refinery_unit', refinery_unit, 'type_equipment', type_equipment,
                   'equipment_raw', equipment_raw, 'status_operation', status_operation,
                   'status_n0', status_n0, 'unit_measurement', unit_measurement,
                   'desain', desain, 'kapasitas_max', kapasitas_max,
                   'average_actual', average_actual, 'remark', remark,
                   'month_update', month_update),
               source_file, source_sheet, source_row, source_record_id
        FROM power_steam_stage WHERE ps_id IS NOT NULL ORDER BY ps_id, source_row
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT equipment_id, ps_id, 'EQUIPMENT_HAS_POWER_STEAM',
               'power_steam', 1.0, 'equipment_exact', false,
               json_object('equipment_raw', equipment_raw),
               source_file, source_sheet, source_row, source_record_id
        FROM power_steam_stage WHERE equipment_id IS NOT NULL AND ps_id IS NOT NULL
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT DISTINCT r.refinery_unit_id, s.ps_id, 'REFINERY_UNIT_HAS_POWER_STEAM',
               'power_steam', 1.0, 'refinery_unit_direct', false,
               json_object('refinery_unit', s.refinery_unit),
               s.source_file, s.source_sheet, s.source_row, s.source_record_id
        FROM power_steam_stage s
        JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
        WHERE s.ps_id IS NOT NULL
    """)


def _build_critical_equipment_nodes(con: duckdb.DuckDBPyConnection, views: list[str]) -> None:
    if not views:
        return
    c = _union_cols(con, views)
    union_sql = " UNION ALL BY NAME ".join(f"SELECT * FROM {v}" for v in views)
    ru_expr = f"ru_normalize({_cs(c, 'refinery_unit','ru','kilang', default='NULL')})"
    eq_expr = _cs(c, 'equipment', 'tag_number', 'tag_no')
    con.execute(f"""
        CREATE TABLE crit_eq_stage AS
        SELECT
            {_cs(c, 'unit_proses','type_equipment','type')} AS unit_proses,
            {eq_expr} AS equipment_raw,
            {_cs(c, 'highlight_issue','problem','masalah')} AS highlight_issue,
            {_cs(c, 'corrective_action','action_plan')} AS corrective_action,
            {_cs(c, 'target_corrective','target_date','target')} AS target_corrective,
            {_cs(c, 'traffic_corrective','traffic')} AS traffic_corrective,
            {_cs(c, 'mitigasi_action','mitigasi')} AS mitigasi_action,
            {_cs(c, 'target_mitigasi')} AS target_mitigasi,
            {_cs(c, 'traffic_mitigasi')} AS traffic_mitigasi,
            {_cs(c, 'month_update')} AS month_update,
            {_cs(c, 'code_current', cast=True, default="'0'")} AS code_current,
            coalesce({ru_expr}, ru_from_filename(_input_source_file)) AS refinery_unit,
            _input_source_file AS source_file,
            _input_source_sheet AS source_sheet,
            _source_row AS source_row,
            'record_' || md5(_input_source_file || '|' || cast(_source_row AS VARCHAR)) AS source_record_id
        FROM ({union_sql})
        WHERE nullif(trim({eq_expr}), '') IS NOT NULL
    """)
    con.execute("""
        ALTER TABLE crit_eq_stage ADD COLUMN ce_id VARCHAR;
        UPDATE crit_eq_stage
        SET ce_id = 'node_critical_equipment_' || md5(source_record_id);
        ALTER TABLE crit_eq_stage ADD COLUMN equipment_id VARCHAR;
        UPDATE crit_eq_stage SET equipment_id = e.equipment_id
        FROM equipment_master e
        WHERE crit_eq_stage.refinery_unit = e.refinery_unit
          AND norm_code(crit_eq_stage.equipment_raw) = norm_code(e.equipment_code_clean);
    """)
    con.execute("""
        INSERT INTO node_raw
        SELECT ce_id, 'critical_equipment', source_record_id,
               coalesce(nullif(equipment_raw,''), 'Critical Equipment'), 'critical_equipment',
               json_object('refinery_unit', refinery_unit, 'unit_proses', unit_proses,
                   'equipment_raw', equipment_raw, 'highlight_issue', highlight_issue,
                   'corrective_action', corrective_action, 'target_corrective', target_corrective,
                   'traffic_corrective', traffic_corrective, 'mitigasi_action', mitigasi_action,
                   'target_mitigasi', target_mitigasi, 'traffic_mitigasi', traffic_mitigasi,
                   'month_update', month_update, 'code_current', code_current),
               source_file, source_sheet, source_row, source_record_id
        FROM crit_eq_stage WHERE ce_id IS NOT NULL
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT equipment_id, ce_id, 'EQUIPMENT_HAS_CRITICAL_ISSUE',
               'critical_equipment', 1.0, 'equipment_exact', false,
               json_object('equipment_raw', equipment_raw),
               source_file, source_sheet, source_row, source_record_id
        FROM crit_eq_stage WHERE equipment_id IS NOT NULL AND ce_id IS NOT NULL
    """)
    con.execute("""
        INSERT INTO relationship_raw (source_node_id, target_node_id, relationship_type,
            domain, confidence, match_rule, is_candidate, properties_json,
            source_file, source_sheet, source_row, source_record_id)
        SELECT DISTINCT r.refinery_unit_id, s.ce_id, 'REFINERY_UNIT_HAS_CRITICAL_EQUIPMENT',
               'critical_equipment', 1.0, 'refinery_unit_direct', false,
               json_object('refinery_unit', s.refinery_unit),
               s.source_file, s.source_sheet, s.source_row, s.source_record_id
        FROM crit_eq_stage s
        JOIN ru_reference r ON s.refinery_unit = r.refinery_unit
        WHERE s.ce_id IS NOT NULL
    """)


# ---------------------------------------------------------------------------
# Output writer
# ---------------------------------------------------------------------------

def _write_outputs(con: duckdb.DuckDBPyConnection, out_dir: Path) -> dict[str, int]:
    out_dir.mkdir(parents=True, exist_ok=True)

    # Deduplikasi node
    con.execute("""
        CREATE TABLE nodes_final AS
        SELECT DISTINCT ON (node_id) node_id, node_type, business_key, label, properties_json
        FROM node_raw ORDER BY node_id, source_row
    """)

    # Deduplikasi relationship (ambil confidence tertinggi)
    con.execute("""
        CREATE TABLE relationships_final AS
        SELECT source_node_id,
               'rel_' || md5(source_node_id || '|' || relationship_type || '|' || target_node_id) AS relationship_id,
               target_node_id, relationship_type, domain, confidence, match_rule, is_candidate, properties_json
        FROM (
            SELECT *, row_number() OVER (
                PARTITION BY source_node_id, relationship_type, target_node_id, is_candidate
                ORDER BY confidence DESC
            ) AS rn
            FROM relationship_raw
            WHERE source_node_id IN (SELECT node_id FROM nodes_final)
              AND target_node_id IN (SELECT node_id FROM nodes_final)
        ) WHERE rn = 1
    """)

    # nodes.csv — format sesuai importer KGRRE
    con.execute(f"""
        COPY (
            SELECT node_id, node_type, label, properties_json
            FROM nodes_final
        ) TO '{out_dir}/nodes.csv' (HEADER, DELIMITER ',')
    """)

    # relationships.csv — verified
    con.execute(f"""
        COPY (
            SELECT relationship_id, source_node_id, target_node_id, relationship_type
            FROM relationships_final WHERE NOT is_candidate
        ) TO '{out_dir}/relationships.csv' (HEADER, DELIMITER ',')
    """)

    # relationship_candidates.csv
    con.execute(f"""
        COPY (
            SELECT relationship_id, source_node_id, target_node_id, relationship_type,
                   confidence, match_rule
            FROM relationships_final WHERE is_candidate
        ) TO '{out_dir}/relationship_candidates.csv' (HEADER, DELIMITER ',')
    """)

    # domain CSVs — dari node_raw dengan properties_json diexpand
    for domain, node_type, fname in [
        ('asset', 'equipment', 'domain_equipment.csv'),
        ('maintenance', 'maintenance_order', 'domain_maintenance.csv'),
        ('reliability', 'reliability_observation', 'domain_reliability.csv'),
        ('inspection_issue', 'inspection', 'domain_inspection_issue.csv'),
        ('readiness_operation', 'readiness_record', 'domain_readiness_operation.csv'),
        ('cost_program', 'rkap_program', 'domain_cost_program.csv'),
        ('oa', 'oa_availability', 'domain_oa_availability.csv'),
        ('oa', 'oa_issue', 'domain_oa_issue.csv'),
        ('plo', 'plo_permit', 'domain_plo_permit.csv'),
    ]:
        con.execute(f"""
            COPY (
                SELECT node_id, node_type, label,
                       source_file, source_sheet, source_row, source_record_id,
                       properties_json
                FROM node_raw WHERE node_type = '{node_type}'
            ) TO '{out_dir}/{fname}' (HEADER, DELIMITER ',')
        """)

    # audit files
    con.execute(f"""
        COPY (SELECT * FROM unmatched_raw) TO '{out_dir}/unmatched_identifier.csv' (HEADER, DELIMITER ',')
    """)

    counts = con.execute("""
        SELECT
            (SELECT count(*) FROM nodes_final),
            (SELECT count(*) FROM relationships_final WHERE NOT is_candidate),
            (SELECT count(*) FROM relationships_final WHERE is_candidate),
            (SELECT count(*) FROM unmatched_raw)
    """).fetchone()

    return {
        "nodes": counts[0], "relationships": counts[1],
        "candidates": counts[2], "unmatched": counts[3]
    }


# ---------------------------------------------------------------------------
# Main ETL runner
# ---------------------------------------------------------------------------

def _run_etl(job: ImportJob, excel_paths: list[Path], out_dir: Path, append: bool = False) -> None:
    try:
        job.status = "running"
        job.phase = "Memuat file Excel"
        job.progress = 5

        con = duckdb.connect()
        con.execute(_MACROS_SQL)
        con.execute(_REFERENCE_SQL)

        # Classify sheets by domain
        domain_views: dict[str, list[str]] = {
            "equipment": [], "rkap": [],
            "reliability": [], "inspection": [], "readiness": [],
            "icu_issue": [], "org_issue": [], "rcps": [], "rcps_recommendation": [],
            "oa_allowance": [], "oa_availability": [], "oa_issue": [], "plo": [],
            "bad_actor": [], "zero_clamp": [], "paf": [], "paf_issue": [],
            "atg": [], "atg_program": [], "pipeline_inspection": [],
            "tkdn": [], "monitoring_operasi": [], "power_steam": [], "critical_equipment": [],
            "work_order": [], "notification": [], "rotor": [],
            "spm_workplan": [], "tank_workplan": [], "jetty_workplan": [],
            "readiness_tank": [], "readiness_jetty": [], "readiness_spm": [],
        }

        job.phase = "Membaca sheet Excel"
        for path in excel_paths:
            job.message = f"Membaca {path.name}…"
            loaded = _load_excel_to_duckdb(con, path)
            for tname, filename, sheet, headers in loaded:
                domain = _detect_domain(filename, sheet, headers)
                if domain and domain in domain_views:
                    domain_views[domain].append(tname)

        build_warnings: list[str] = []

        def _safe(phase: str, fn, *args):
            try:
                fn(*args)
            except Exception as exc:
                build_warnings.append(f"{phase}: {exc}")

        job.progress = 20
        job.phase = "Membangun node RU & Plant"
        _build_ru_plant_nodes(con)

        job.progress = 25
        job.phase = "Membangun node Equipment"
        _safe("Equipment", _build_equipment_nodes, con, domain_views["equipment"])
        # Pastikan equipment_master ada agar builder lain tidak crash jika equipment builder gagal
        con.execute("""
            CREATE TABLE IF NOT EXISTS equipment_master (
                equipment_id VARCHAR, equipment_code_normalized VARCHAR,
                equipment_code_raw VARCHAR, equipment_code_clean VARCHAR, refinery_unit VARCHAR,
                plant VARCHAR, functional_location VARCHAR,
                equipment_group VARCHAR, equip_category VARCHAR,
                description VARCHAR, criticallity VARCHAR,
                plant_area VARCHAR, manufacturer VARCHAR,
                model_type VARCHAR, wbs_element VARCHAR,
                cost_center VARCHAR, planner_group VARCHAR,
                date_update_data VARCHAR, source_file VARCHAR,
                source_sheet VARCHAR, source_row INTEGER,
                source_record_id VARCHAR, rn INTEGER
            )
        """)

        job.progress = 45
        job.phase = "Membangun node RKAP Program"
        _safe("RKAP", _build_rkap_nodes, con, domain_views["rkap"])

        job.progress = 55
        job.phase = "Membangun node Reliability"
        _safe("Reliability", _build_reliability_nodes, con, domain_views["reliability"])

        job.progress = 62
        job.phase = "Membangun node Inspection"
        _safe("Inspection", _build_inspection_nodes, con, domain_views["inspection"])

        job.progress = 68
        job.phase = "Membangun node ICU/Org Issue"
        _safe("ICU Issue", _build_icu_issue_nodes, con, domain_views["icu_issue"] + domain_views["org_issue"])

        job.progress = 74
        job.phase = "Membangun node Readiness"
        _safe("Readiness", _build_readiness_nodes, con, domain_views["readiness"])

        job.progress = 76
        job.phase = "Membangun node OA Data"
        _safe("OA Data", _build_oa_nodes, con, domain_views["oa_allowance"], domain_views["oa_availability"], domain_views["oa_issue"])

        job.progress = 79
        job.phase = "Membangun node PLO"
        _safe("PLO", _build_plo_nodes, con, domain_views["plo"])

        job.progress = 80
        job.phase = "Membangun node RCPS"
        _safe("RCPS", _build_rcps_nodes, con, domain_views["rcps"], domain_views["rcps_recommendation"])

        job.progress = 81
        job.phase = "Membangun node Bad Actor"
        _safe("Bad Actor", _build_bad_actor_nodes, con, domain_views["bad_actor"])

        job.progress = 82
        job.phase = "Membangun node Zero Clamp"
        _safe("Zero Clamp", _build_zero_clamp_nodes, con, domain_views["zero_clamp"])

        job.progress = 83
        job.phase = "Membangun node PAF"
        _safe("PAF", _build_paf_nodes, con, domain_views["paf"], domain_views["paf_issue"])

        job.progress = 84
        job.phase = "Membangun node ATG"
        _safe("ATG", _build_atg_nodes, con, domain_views["atg"], domain_views["atg_program"])

        job.progress = 85
        job.phase = "Membangun node Pipeline Inspection"
        _safe("Pipeline Inspection", _build_pipeline_inspection_nodes, con, domain_views["pipeline_inspection"])

        job.progress = 86
        job.phase = "Membangun node TKDN"
        _safe("TKDN", _build_tkdn_nodes, con, domain_views["tkdn"])

        job.progress = 87
        job.phase = "Membangun node Monitoring Operasi"
        _safe("Monitoring Operasi", _build_monitoring_operasi_nodes, con, domain_views["monitoring_operasi"])

        job.progress = 88
        job.phase = "Membangun node Power & Steam"
        _safe("Power Steam", _build_power_steam_nodes, con, domain_views["power_steam"])

        job.progress = 89
        job.phase = "Membangun node Critical Equipment"
        _safe("Critical Equipment", _build_critical_equipment_nodes, con, domain_views["critical_equipment"])

        job.progress = 89
        job.phase = "Membangun node Rotor"
        _safe("Rotor", _build_rotor_nodes, con, domain_views["rotor"])

        job.progress = 89
        job.phase = "Membangun node Work Order"
        _safe("Work Order", _build_work_order_nodes, con, domain_views["work_order"])

        job.progress = 89
        job.phase = "Membangun node Notifikasi SAP"
        _safe("Notification", _build_notification_nodes, con, domain_views["notification"])

        job.progress = 89
        job.phase = "Membangun node Program Kerja SPM"
        _safe("SPM Workplan", _build_workplan_nodes, con, domain_views["spm_workplan"],
              "spm_workplan", "spm_workplan", "SPM", "spm_workplan_stage",
              "EQUIPMENT_HAS_SPM_WORKPLAN", "REFINERY_UNIT_HAS_SPM_WORKPLAN")

        job.progress = 89
        job.phase = "Membangun node Program Kerja Tank"
        _safe("Tank Workplan", _build_workplan_nodes, con, domain_views["tank_workplan"],
              "tank_workplan", "tank_workplan", "Tank", "tank_workplan_stage",
              "EQUIPMENT_HAS_TANK_WORKPLAN", "REFINERY_UNIT_HAS_TANK_WORKPLAN")

        job.progress = 89
        job.phase = "Membangun node Program Kerja Jetty"
        _safe("Jetty Workplan", _build_workplan_nodes, con, domain_views["jetty_workplan"],
              "jetty_workplan", "jetty_workplan", "Jetty", "jetty_workplan_stage",
              "EQUIPMENT_HAS_JETTY_WORKPLAN", "REFINERY_UNIT_HAS_JETTY_WORKPLAN")

        job.progress = 89
        job.phase = "Membangun node Readiness Tank"
        _safe("Readiness Tank", _build_readiness_subtype_nodes, con, domain_views["readiness_tank"],
              "readiness_tank", "Tank", "readiness_tank_stage", [])

        job.progress = 89
        job.phase = "Membangun node Readiness Jetty"
        _safe("Readiness Jetty", _build_readiness_subtype_nodes, con, domain_views["readiness_jetty"],
              "readiness_jetty", "Jetty", "readiness_jetty_stage", [])

        job.progress = 89
        job.phase = "Membangun node Readiness SPM"
        _safe("Readiness SPM", _build_readiness_subtype_nodes, con, domain_views["readiness_spm"],
              "readiness_spm", "SPM", "readiness_spm_stage", [])

        job.progress = 90
        job.phase = "Membangun relasi antar domain"
        _safe("Cross-domain", _build_cross_domain_relationships, con)

        job.progress = 91
        job.phase = "Menulis output CSV"
        counts = _write_outputs(con, out_dir)
        con.close()

        job.progress = 88
        job.phase = "Import ke database"
        job.warnings = build_warnings
        job.message = (
            f"ETL selesai: {counts['nodes']:,} node, {counts['relationships']:,} relasi, "
            f"{counts['candidates']:,} kandidat, {counts['unmatched']:,} tidak cocok"
        )

        # Trigger existing import pipeline
        from .scanner import scan_package
        from .importer import _select_ready_files, _run_import
        scan = scan_package(out_dir, validate=True)
        files = _select_ready_files(scan, allow_partial=True)
        _run_import(job, files, out_dir, True, existing_dataset_id=job.dataset_id, append=append)

    except Exception as exc:
        job.status = "failed"
        job.error = str(exc)
        job.message = str(exc)
        job.finished_at = time.time()


def start_etl_import(name: str, excel_paths: list[Path], existing_dataset_id: str | None = None,
                     append: bool = False) -> ImportJob:
    out_dir = UPLOADS_DIR / uuid.uuid4().hex
    out_dir.mkdir(parents=True, exist_ok=True)
    job = _create_job(name)
    job.dataset_id = existing_dataset_id  # pre-set agar _run_import tahu ini sync
    with ETL_JOBS_LOCK:
        ETL_JOBS[job.id] = job
    threading.Thread(
        target=_run_etl, args=(job, excel_paths, out_dir, append), daemon=True
    ).start()
    return job
